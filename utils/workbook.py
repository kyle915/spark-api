"""Render a tabular export payload to XLSX / CSV.

Deliberately DJANGO-FREE: it takes the plain dict that
:func:`recaps.recap_field_export.build_recap_field_export` produces and needs
nothing but ``openpyxl`` and the stdlib. That lets a GitHub Actions runner turn
the JSON artifact into a client-ready workbook without a database, credentials,
or a Django settings module (see ``scripts/build_recap_workbook.py``), while the
management command can call the exact same code in-process.

The payload contract is small on purpose:

    {
      "columns": [{"header": str, "group": str, "kind": "text"|"number"|"links"}],
      "rows":    [[cell, ...]],          # positionally aligned to columns
      "files":   [{"category","name","url","recap_uuid","event_name",...}],
      "tenant":  {"name": str, ...},
      "window":  {"start": str|None, "end": str|None},
      "meta":    {...},
      "diagnostics": {...},
    }

A ``links`` cell is a list of URL strings. Excel allows exactly ONE hyperlink
per cell, so a multi-file cell cannot be made clickable per-URL — those cells
carry the URLs as wrapped, copy-pasteable text and a dedicated
"Photos & Receipts" sheet holds one clickable row per file. Single-URL cells do
get a real hyperlink, since that costs nothing.
"""
from __future__ import annotations

import csv
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATA_SHEET = "Recap Data"
FILES_SHEET = "Photos & Receipts"
NOTES_SHEET = "Read Me"

# openpyxl refuses control characters; strip them rather than blow up an
# export because one BA pasted a stray \x0b into a free-text answer.
_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Excel's hard per-cell ceiling.
_CELL_MAX = 32767

_HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_GROUP_FILL = PatternFill("solid", fgColor="E8EDF2")
_GROUP_FONT = Font(bold=True, color="1F2A37", size=10)
_TITLE_FONT = Font(bold=True, size=14)
_NOTE_FONT = Font(size=10)
_LINK_FONT = Font(color="0B5CAB", underline="single", size=10)
_THIN = Side(style="thin", color="D0D7DE")
_HEADER_BORDER = Border(bottom=_THIN)


def _clean(value) -> str:
    """Excel-safe string for any scalar."""
    if value is None:
        return ""
    text = str(value)
    text = _ILLEGAL_RE.sub("", text)
    if len(text) > _CELL_MAX:
        text = text[: _CELL_MAX - 1] + "…"
    return text


def _cell_text(value, kind: str) -> str:
    """Flatten one payload cell to the text that lands in the sheet."""
    if kind == "links":
        urls = value or []
        if isinstance(urls, str):
            urls = [urls]
        return _clean("\n".join(u for u in urls if u))
    return _clean(value)


def _autosize(ws, headers: list[str], rows: list[list], *, link_cols: set[int]) -> None:
    """Width per column from a sample of its content.

    Link columns get a fixed, generous-but-bounded width — a full GCS URL is
    ~150 chars and letting that drive the width would make the sheet unusable.
    """
    sample = rows[:400]
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        if idx - 1 in link_cols:
            ws.column_dimensions[letter].width = 46
            continue
        widest = len(header)
        for row in sample:
            if idx - 1 >= len(row):
                continue
            text = row[idx - 1]
            if not text:
                continue
            longest_line = max((len(part) for part in str(text).split("\n")), default=0)
            widest = max(widest, longest_line)
        ws.column_dimensions[letter].width = max(10, min(widest + 2, 52))


def _write_data_sheet(wb: Workbook, payload: dict) -> None:
    columns = payload.get("columns") or []
    rows = payload.get("rows") or []
    headers = [c.get("header", "") for c in columns]
    kinds = [c.get("kind", "text") for c in columns]
    groups = [c.get("group", "") for c in columns]
    link_cols = {i for i, k in enumerate(kinds) if k == "links"}

    ws = wb.create_sheet(DATA_SHEET)

    # Row 1: the column GROUP (Identity / each template section / Files), so a
    # reader can see which template section a field came from. The label is
    # written once at the FIRST column of each run — repeating "Identity" 14
    # times across the band is noise. Row 2: the headers themselves.
    for idx, group in enumerate(groups, start=1):
        starts_run = idx == 1 or groups[idx - 2] != group
        cell = ws.cell(row=1, column=idx, value=_clean(group) if starts_run else None)
        cell.fill = _GROUP_FILL
        cell.font = _GROUP_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=idx, value=_clean(header))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    text_rows: list[list[str]] = []
    for r_i, row in enumerate(rows, start=3):
        flat: list[str] = []
        for c_i, kind in enumerate(kinds):
            raw = row[c_i] if c_i < len(row) else ""
            if kind == "number":
                cell = ws.cell(row=r_i, column=c_i + 1, value=raw if raw != "" else None)
                flat.append(_clean(raw))
                continue
            text = _cell_text(raw, kind)
            cell = ws.cell(row=r_i, column=c_i + 1, value=text)
            if kind == "links":
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                urls = raw if isinstance(raw, list) else ([raw] if raw else [])
                # One hyperlink per cell is all Excel allows — wire it up when
                # the cell holds exactly one file; multi-file cells stay text
                # and are clickable on the "Photos & Receipts" sheet.
                if len(urls) == 1:
                    cell.hyperlink = urls[0]
                    cell.font = _LINK_FONT
            else:
                cell.alignment = Alignment(vertical="top")
            flat.append(text)
        text_rows.append(flat)

    ws.freeze_panes = "A3"
    if headers:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(len(rows) + 2, 3)}"
    _autosize(ws, headers, text_rows, link_cols=link_cols)


def _write_files_sheet(wb: Workbook, payload: dict) -> None:
    """One clickable row per file — the tab a client actually clicks through."""
    files = payload.get("files") or []
    ws = wb.create_sheet(FILES_SHEET)
    headers = [
        "Category",
        "Event",
        "Event Date",
        "Brand Ambassador",
        "File Name",
        "Link",
        "Recap ID",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER

    for r_i, f in enumerate(files, start=2):
        ws.cell(row=r_i, column=1, value=_clean(f.get("category")))
        ws.cell(row=r_i, column=2, value=_clean(f.get("event_name")))
        ws.cell(row=r_i, column=3, value=_clean(f.get("event_date")))
        ws.cell(row=r_i, column=4, value=_clean(f.get("ba")))
        ws.cell(row=r_i, column=5, value=_clean(f.get("name")))
        url = (f.get("url") or "").strip()
        link = ws.cell(row=r_i, column=6, value=_clean(url) or None)
        if url:
            link.hyperlink = url
            link.font = _LINK_FONT
        ws.cell(row=r_i, column=7, value=_clean(f.get("recap_uuid")))

    ws.freeze_panes = "A2"
    if files:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(files) + 1}"
    for idx, width in enumerate((22, 34, 12, 22, 34, 62, 34), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _notes_lines(payload: dict) -> list[tuple[str, str]]:
    """(label, value) pairs for the Read Me sheet.

    This sheet exists to stop two specific misreadings that have bitten this
    data before: treating a blank as a zero, and treating "samples distributed"
    and "consumers reached" as the same number.
    """
    tenant = payload.get("tenant") or {}
    meta = payload.get("meta") or {}
    window = payload.get("window") or {}
    diag = payload.get("diagnostics") or {}
    start, end = window.get("start"), window.get("end")
    window_label = (
        f"{start or 'earliest'} → {end or 'latest'} (by EVENT date)"
        if (start or end)
        else "All time (no date filter)"
    )
    rows: list[tuple[str, str]] = [
        ("Tenant", str(tenant.get("name") or "")),
        ("Date window", window_label),
        ("Generated", str(payload.get("generated_at") or "")),
        ("", ""),
        ("GRAIN", "One row per RECAP — not per event."),
        (
            "Rows",
            f"{meta.get('row_count', 0)} recaps. Events in this window with no recap "
            f"filed: {diag.get('events_without_recap', 0)} — those are NOT rows here, "
            f"because a blank row would read as a measured zero.",
        ),
        (
            "Columns",
            f"{meta.get('column_count', 0)} total — "
            f"{meta.get('identity_column_count', 0)} identity, "
            f"{meta.get('field_column_count', 0)} template fields, "
            f"{meta.get('file_column_count', 0)} file columns.",
        ),
        ("Files", f"{meta.get('file_count', 0)} photos/receipts linked."),
        ("", ""),
        (
            "READ THIS — blanks",
            "A blank cell means the question was NOT ANSWERED, or the form does not "
            "ask it. It does not mean zero. Metrics this tenant's form never "
            "collects have no column at all, on purpose.",
        ),
        (
            "READ THIS — samples vs consumers",
            "Samples distributed and consumers reached are DIFFERENT measures and do "
            "not reconcile to each other. Each stays under its own field name here; "
            "never relabel one as the other.",
        ),
        (
            "Links",
            "Every file link is a permanent, unsigned public URL — it does not "
            "expire. Multi-file cells list URLs as text; the "
            f"'{FILES_SHEET}' tab has one clickable row per file.",
        ),
    ]
    by_cat = diag.get("files_by_category") or {}
    if by_cat:
        rows.append(("", ""))
        rows.append(("Files by category", ""))
        for name, count in sorted(by_cat.items(), key=lambda kv: (-kv[1], kv[0])):
            rows.append((f"    {name}", str(count)))

    over = diag.get("consumers_exceeding_engagements") or {}
    if over.get("count"):
        rows.append(("", ""))
        rows.append(
            (
                "Data to verify",
                f"{over['count']} recap(s) report more consumers sampled than total "
                f"engagements, which is not possible — together overstating consumers "
                f"by {over.get('total_excess', 0)}. Those rows are included as filed; "
                f"treat their consumer counts as unconfirmed.",
            )
        )

    review = (diag.get("unapproved_or_draft_rows") or 0) + (
        diag.get("internal_demo_rows") or 0
    )
    if review:
        rows.append(("", ""))
        rows.append(
            (
                "Rows to review",
                f"{diag.get('unapproved_or_draft_rows', 0)} unapproved/draft and "
                f"{diag.get('internal_demo_rows', 0)} internal-demo row(s) are included. "
                f"Nothing was dropped automatically — remove them if this workbook is "
                f"going to a client as-is.",
            )
        )

    unlinkable = diag.get("files_without_a_link") or 0
    if unlinkable:
        rows.append(("", ""))
        rows.append(
            (
                "Missing links",
                f"{unlinkable} attached file(s) could not be turned into a link and are "
                f"NOT represented above. The files exist — treat this workbook's file "
                f"columns as incomplete until that is resolved.",
            )
        )

    misfiled = (diag.get("receipt_looking_files_outside_receipt_categories") or {}).get("count", 0)
    if misfiled:
        rows.append(("", ""))
        rows.append(
            (
                "Category warning",
                f"{misfiled} file(s) look like receipts by filename but sit in a "
                f"non-receipt category. Check before treating a photo column as "
                f"purely photos.",
            )
        )
    empties = diag.get("field_columns_entirely_empty") or []
    if empties:
        rows.append(("", ""))
        rows.append(
            (
                "Empty field columns",
                f"{len(empties)} template field(s) had no answer in ANY row: "
                + ", ".join(empties[:25])
                + ("…" if len(empties) > 25 else ""),
            )
        )
    return rows


def _write_notes_sheet(wb: Workbook, payload: dict) -> None:
    ws = wb.create_sheet(NOTES_SHEET)
    title = ws.cell(row=1, column=1, value="Field-Level Recap Export")
    title.font = _TITLE_FONT
    for r_i, (label, value) in enumerate(_notes_lines(payload), start=3):
        lcell = ws.cell(row=r_i, column=1, value=_clean(label))
        lcell.font = Font(bold=True, size=10)
        lcell.alignment = Alignment(vertical="top")
        vcell = ws.cell(row=r_i, column=2, value=_clean(value))
        vcell.font = _NOTE_FONT
        vcell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 104


def build_xlsx(payload: dict) -> bytes:
    """Render the payload to XLSX bytes: Read Me, Recap Data, Photos & Receipts."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    _write_notes_sheet(wb, payload)
    _write_data_sheet(wb, payload)
    _write_files_sheet(wb, payload)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_grid_csv(payload: dict) -> str:
    """The main grid as CSV. Link cells become newline-joined URLs (quoted)."""
    columns = payload.get("columns") or []
    kinds = [c.get("kind", "text") for c in columns]
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow([c.get("header", "") for c in columns])
    for row in payload.get("rows") or []:
        writer.writerow(
            [_cell_text(row[i] if i < len(row) else "", kinds[i]) for i in range(len(columns))]
        )
    return buf.getvalue()


def build_files_csv(payload: dict) -> str:
    """The per-file list as CSV — one row per photo/receipt."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(
        ["Category", "Event", "Event Date", "Brand Ambassador", "File Name", "Link", "Recap ID"]
    )
    for f in payload.get("files") or []:
        writer.writerow(
            [
                f.get("category", ""),
                f.get("event_name", ""),
                f.get("event_date", ""),
                f.get("ba", ""),
                f.get("name", ""),
                f.get("url", ""),
                f.get("recap_uuid", ""),
            ]
        )
    return buf.getvalue()


def write_outputs(payload: dict, base_path: str) -> dict:
    """Write ``<base>.xlsx``, ``<base>.csv`` and ``<base>-files.csv``.

    ``base_path`` may carry a ``.xlsx`` suffix (it's stripped) so callers can
    pass either a bare stem or the intended workbook path.
    """
    base = base_path[:-5] if base_path.lower().endswith(".xlsx") else base_path
    xlsx_path = f"{base}.xlsx"
    csv_path = f"{base}.csv"
    files_csv_path = f"{base}-files.csv"

    with open(xlsx_path, "wb") as fh:
        fh.write(build_xlsx(payload))
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        fh.write(build_grid_csv(payload))
    with open(files_csv_path, "w", newline="", encoding="utf-8") as fh:
        fh.write(build_files_csv(payload))

    return {"xlsx": xlsx_path, "csv": csv_path, "files_csv": files_csv_path}
