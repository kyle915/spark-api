from __future__ import annotations

from dataclasses import dataclass
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Any
import datetime

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import Workbook, load_workbook

from events.models import (
    Client,
    Distributor,
    Event,
    EventStatus,
    EventType,
    Location,
    Product,
    Request,
    RequestDetail,
    RequestProduct,
    RequestStatus,
    RequestStoreManager,
    RequestType,
    Retailer,
    Tenant,
    TimeZone,
)

User = get_user_model()

TEMPLATE_COLUMNS = [
    "name",
    "date",
    "start_time",
    "end_time",
    "address",
    "latitude",
    "longitude",
    "notes",
    "requestor_email",
    "distributor_name",
    "retailer_name",
    "city",
    "store_manager_name",
    "store_manager_phone",
    "timezone_code",
    "request_type_id",
    "event_type_id",
    "store_manager_id",
    "table_size",
    "product_ids",
]

REQUIRED_COLUMNS = [
    "name",
    "date",
    "start_time",
    "end_time",
    "address",
]


@dataclass
class BatchRequestRowResult:
    row_number: int
    success: bool
    message: str
    request_id: int | None = None
    request_uuid: str | None = None


@dataclass
class BatchRequestImportResult:
    total_rows: int
    success_count: int
    failed_count: int
    rows: list[BatchRequestRowResult]
    rolled_back: bool = False


def export_request_batch_template(output_path: str) -> Path:
    template_bytes = build_request_batch_template_xlsx()
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(template_bytes)
    return path


def build_request_batch_template_xlsx(tenant_id: int | None = None) -> bytes:
    sample = {
        "name": "North Point Activation",
        "date": "02/20/2026",
        "start_time": "10:00",
        "end_time": "14:00",
        "address": "123 Main St",
        "latitude": 40.7128,
        "longitude": -74.006,
        "notes": "Product demo and sampling",
        "requestor_email": "requestor@example.com",
        "distributor_name": "North Distribution",
        "retailer_name": "Central Retail",
        "city": "New York",
        "store_manager_name": "John Store Manager",
        "store_manager_phone": "+1-555-0101",
        "timezone_code": "EST",
        "request_type_id": 1,
        "event_type_id": 1,
        "store_manager_id": "",
        "table_size": 1,
        "product_ids": "3,7,10",
    }

    wb = Workbook()

    ws_requests = wb.active
    ws_requests.title = "Requests"
    ws_requests.append(TEMPLATE_COLUMNS)
    ws_requests.append([sample.get(col) for col in TEMPLATE_COLUMNS])

    ws_timezones = wb.create_sheet("TimeZones")
    ws_timezones.append(["id", "name", "code"])
    for row in TimeZone.objects.all().order_by("id").values_list("id", "name", "code"):
        ws_timezones.append(list(row))

    ws_request_types = wb.create_sheet("RequestTypes")
    ws_request_types.append(["id", "name"])
    request_type_qs = RequestType.objects.all()
    if tenant_id:
        request_type_qs = request_type_qs.filter(tenant_id=tenant_id)
    for row in request_type_qs.order_by("id").values_list("id", "name"):
        ws_request_types.append(list(row))

    ws_event_types = wb.create_sheet("EventTypes")
    ws_event_types.append(["id", "name"])
    event_type_qs = EventType.objects.all()
    if tenant_id:
        event_type_qs = event_type_qs.filter(tenant_id=tenant_id)
    for row in event_type_qs.order_by("id").values_list("id", "name"):
        ws_event_types.append(list(row))

    ws_cities = wb.create_sheet("Cities")
    ws_cities.append(["id", "name", "code", "zip", "state_id", "state_name"])
    location_qs = Location.objects.select_related("state").all()
    if tenant_id:
        location_qs = location_qs.filter(tenant_id=tenant_id)
    for row in location_qs.order_by("id").values(
        "id", "name", "code", "zip", "state_id", "state__name"
    ):
        ws_cities.append(
            [
                row["id"],
                row["name"],
                row["code"],
                row["zip"],
                row["state_id"],
                row["state__name"],
            ]
        )

    ws_products = wb.create_sheet("Products")
    ws_products.append(["id", "name", "product_type"])
    product_qs = Product.objects.all()
    if tenant_id:
        product_qs = product_qs.filter(tenant_id=tenant_id)
    for row in product_qs.order_by("id").values("id", "name", "product_type__name"):
        ws_products.append([row["id"], row["name"], row["product_type__name"]])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def import_requests_from_excel(
    *,
    file_path: str,
    tenant_id: int,
    created_by_id: int,
    default_timezone_id: int | None = None,
    default_request_type_id: int | None = None,
    sheet_name: str | int = 0,
    dry_run: bool = False,
    rollback_on_error: bool = True,
) -> BatchRequestImportResult:
    if not Path(file_path).exists():
        raise ValueError(f"File not found: {file_path}")

    tenant = Tenant.objects.filter(id=tenant_id).first()
    if not tenant:
        raise ValueError(f"Tenant not found: {tenant_id}")

    created_by = User.objects.filter(id=created_by_id).first()
    if not created_by:
        raise ValueError(f"User not found: {created_by_id}")

    approved_status = RequestStatus.objects.get_by_slug("approved", tenant=tenant_id)
    if not approved_status:
        raise ValueError(
            "Approved request status not found for tenant. Create status with slug 'approved'."
        )

    if default_timezone_id and not TimeZone.objects.filter(id=default_timezone_id).exists():
        raise ValueError(f"default_timezone_id does not exist: {default_timezone_id}")

    if default_request_type_id and not RequestType.objects.filter(
        id=default_request_type_id, tenant_id=tenant_id
    ).exists():
        raise ValueError(
            "default_request_type_id does not exist for this tenant: "
            f"{default_request_type_id}"
        )

    wb = load_workbook(filename=file_path, data_only=True)
    headers, rows = _read_rows_from_workbook(wb, sheet_name)

    return _import_requests_from_rows(
        headers=headers,
        rows=rows,
        tenant_id=tenant_id,
        tenant_name=tenant.name,
        created_by=created_by,
        approved_status=approved_status,
        default_timezone_id=default_timezone_id,
        default_request_type_id=default_request_type_id,
        dry_run=dry_run,
        rollback_on_error=rollback_on_error,
    )


def import_requests_from_excel_bytes(
    *,
    file_bytes: bytes,
    tenant_id: int,
    created_by_id: int,
    default_timezone_id: int | None = None,
    default_request_type_id: int | None = None,
    sheet_name: str | int = 0,
    dry_run: bool = False,
    rollback_on_error: bool = True,
) -> BatchRequestImportResult:
    tenant = Tenant.objects.filter(id=tenant_id).first()
    if not tenant:
        raise ValueError(f"Tenant not found: {tenant_id}")

    created_by = User.objects.filter(id=created_by_id).first()
    if not created_by:
        raise ValueError(f"User not found: {created_by_id}")

    approved_status = RequestStatus.objects.get_by_slug("approved", tenant=tenant_id)
    if not approved_status:
        raise ValueError(
            "Approved request status not found for tenant. Create status with slug 'approved'."
        )

    if default_timezone_id and not TimeZone.objects.filter(id=default_timezone_id).exists():
        raise ValueError(f"default_timezone_id does not exist: {default_timezone_id}")

    if default_request_type_id and not RequestType.objects.filter(
        id=default_request_type_id, tenant_id=tenant_id
    ).exists():
        raise ValueError(
            "default_request_type_id does not exist for this tenant: "
            f"{default_request_type_id}"
        )

    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    headers, rows = _read_rows_from_workbook(wb, sheet_name)

    return _import_requests_from_rows(
        headers=headers,
        rows=rows,
        tenant_id=tenant_id,
        tenant_name=tenant.name,
        created_by=created_by,
        approved_status=approved_status,
        default_timezone_id=default_timezone_id,
        default_request_type_id=default_request_type_id,
        dry_run=dry_run,
        rollback_on_error=rollback_on_error,
    )


def _read_rows_from_workbook(
    wb,
    sheet_name: str | int,
) -> tuple[list[str], list[dict[str, Any]]]:
    if isinstance(sheet_name, int):
        if sheet_name < 0 or sheet_name >= len(wb.worksheets):
            raise ValueError(f"sheet_name index out of range: {sheet_name}")
        ws = wb.worksheets[sheet_name]
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"sheet_name not found: {sheet_name}")
        ws = wb[sheet_name]

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        raise ValueError("Excel sheet has no header row.")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_cells]
    if all(not h for h in headers):
        raise ValueError("Excel sheet header is empty.")

    rows: list[dict[str, Any]] = []
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        if row_values is None:
            continue
        if all(_is_empty(value) for value in row_values):
            continue
        row_dict: dict[str, Any] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = row_values[i] if i < len(row_values) else None
        rows.append(row_dict)

    return headers, rows


def _import_requests_from_rows(
    *,
    headers: list[str],
    rows: list[dict[str, Any]],
    tenant_id: int,
    tenant_name: str,
    created_by: User,
    approved_status: RequestStatus,
    default_timezone_id: int | None,
    default_request_type_id: int | None,
    dry_run: bool,
    rollback_on_error: bool,
) -> BatchRequestImportResult:
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing_columns:
        raise ValueError("Missing required columns: " + ", ".join(sorted(missing_columns)))

    if not default_timezone_id and "timezone_code" not in headers:
        raise ValueError(
            "Column 'timezone_code' is required when default_timezone_id is not provided."
        )

    if not default_request_type_id and "request_type_id" not in headers:
        raise ValueError(
            "Column 'request_type_id' is required when default_request_type_id is not provided."
        )

    if "event_type_id" not in headers:
        raise ValueError("Column 'event_type_id' is required.")

    approved_event_status = EventStatus.objects.filter(
        slug="approved",
        tenant_id=tenant_id,
    ).first()
    if not approved_event_status:
        raise ValueError(
            "Approved event status not found for tenant. Create status with slug 'approved'."
        )

    results: list[BatchRequestRowResult] = []
    success_count = 0
    failed_count = 0

    def _process() -> None:
        nonlocal success_count, failed_count, results

        for idx, row in enumerate(rows):
            row_number = idx + 2
            try:
                parsed = _parse_row(
                    row=row,
                    tenant_id=tenant_id,
                    tenant_name=tenant_name,
                    default_timezone_id=default_timezone_id,
                    default_request_type_id=default_request_type_id,
                )

                request = Request(
                    tenant_id=tenant_id,
                    created_by=created_by,
                    status=approved_status,
                    name=parsed["name"],
                    date=parsed["date"],
                    start_time=parsed["start_time"],
                    end_time=parsed["end_time"],
                    address=parsed["address"],
                    notes=parsed["notes"],
                    coordinates=parsed["coordinates_request"],
                    requestor_email=parsed["requestor_email"],
                    client_name=parsed["client_name"],
                    client_email=parsed["client_email"],
                    distributor_name=parsed["distributor_name"],
                    distributor_email=parsed["distributor_email"],
                    retailer_name=parsed["retailer_name"],
                    retailer_address=parsed["retailer_address"],
                    retailer_store_contact=parsed["retailer_store_contact"],
                    store_manager_name=parsed["store_manager_name"],
                    store_manager_phone=parsed["store_manager_phone"],
                    timezone_id=parsed["timezone_id"],
                    request_type_id=parsed["request_type_id"],
                    client_id=parsed["client_id"],
                    distributor_id=parsed["distributor_id"],
                    retailer_id=parsed["retailer_id"],
                )

                if not dry_run:
                    request.save()

                    if parsed["store_manager_id"]:
                        manager = RequestStoreManager.objects.get(id=parsed["store_manager_id"])
                        if manager.tenant_id and manager.tenant_id != tenant_id:
                            raise ValueError("store_manager_id belongs to another tenant.")
                        manager.request = request
                        if not manager.tenant_id:
                            manager.tenant_id = tenant_id
                        manager.updated_by = created_by
                        manager.save()

                    if parsed["create_detail"]:
                        RequestDetail.objects.create(
                            request=request,
                            tenant_id=tenant_id,
                            created_by=created_by,
                            is_table_needed=parsed["is_table_needed"],
                            table_size=parsed["table_size"],
                        )

                    for product_id in parsed["product_ids"]:
                        RequestProduct.objects.create(
                            request=request,
                            product_id=product_id,
                            tenant_id=tenant_id,
                            created_by=created_by,
                        )

                    Event.objects.create(
                        tenant_id=tenant_id,
                        request=request,
                        event_type_id=parsed["event_type_id"],
                        status=approved_event_status,
                        timezone_id=parsed["timezone_id"],
                        name=parsed["name"],
                        date=parsed["date"],
                        start_time=parsed["start_time"],
                        end_time=parsed["end_time"],
                        address=parsed["address"],
                        notes=parsed["notes"],
                        coordinates=parsed["coordinates_event"],
                        retailer_id=parsed["retailer_id"],
                        distributor_id=parsed["distributor_id"],
                        created_by=created_by,
                    )

                success_count += 1
                results.append(
                    BatchRequestRowResult(
                        row_number=row_number,
                        success=True,
                        message="Validated (dry-run)." if dry_run else "Imported.",
                        request_id=request.id if not dry_run else None,
                        request_uuid=str(request.uuid) if not dry_run else None,
                    )
                )
            except Exception as exc:
                failed_count += 1
                results.append(
                    BatchRequestRowResult(
                        row_number=row_number,
                        success=False,
                        message=str(exc),
                    )
                )

    rolled_back = False
    if dry_run or not rollback_on_error:
        _process()
    else:
        with transaction.atomic():
            _process()
            if failed_count > 0:
                rolled_back = True
                transaction.set_rollback(True)

    if rolled_back:
        success_count = 0
        failed_count = len(results)
        for row in results:
            if row.success:
                row.success = False
                row.request_id = None
                row.request_uuid = None
                row.message = "Rolled back because another row failed."

    return BatchRequestImportResult(
        total_rows=len(rows),
        success_count=success_count,
        failed_count=failed_count,
        rows=results,
        rolled_back=rolled_back,
    )


def _parse_row(
    *,
    row: dict[str, Any],
    tenant_id: int,
    tenant_name: str,
    default_timezone_id: int | None,
    default_request_type_id: int | None,
) -> dict[str, Any]:
    errors: list[str] = []

    def _capture(fn, *args, **kwargs):
        try:
            return fn(*args, **kwargs)
        except ValueError as exc:
            errors.append(str(exc))
            return None

    name = _capture(_required_str, row, "name")
    event_date = _capture(_required_date, row, "date")
    start_clock = _capture(_required_time, row, "start_time")
    end_clock = _capture(_required_time, row, "end_time")
    address = _capture(_required_str, row, "address")
    latitude = _capture(_optional_float, row.get("latitude"), "latitude")
    longitude = _capture(_optional_float, row.get("longitude"), "longitude")

    if (latitude is None) != (longitude is None):
        errors.append("latitude and longitude must be provided together.")

    timezone_code = _optional_str(row.get("timezone_code"))
    request_type_id = _capture(_optional_int, row.get("request_type_id"), "request_type_id")
    event_type_id = _capture(_optional_int, row.get("event_type_id"), "event_type_id")

    timezone_id: int | None = None
    timezone_obj: TimeZone | None = None

    if timezone_code:
        timezone = TimeZone.objects.filter(code__iexact=timezone_code).order_by("id").first()
        if not timezone:
            errors.append(f"timezone_code does not exist: {timezone_code}")
        else:
            timezone_obj = timezone
            timezone_id = timezone.id
    else:
        timezone_id = default_timezone_id

    request_type_id = request_type_id or default_request_type_id

    if not timezone_id:
        errors.append("timezone_code is required (or provide default_timezone_id).")
    if not request_type_id:
        errors.append("request_type_id is required.")
    if not event_type_id:
        errors.append("event_type_id is required.")

    if timezone_obj is None and timezone_id:
        timezone_obj = TimeZone.objects.filter(id=timezone_id).first()
    if timezone_id and not timezone_obj:
        errors.append(f"timezone_id does not exist: {timezone_id}")

    date = None
    start_time = None
    end_time = None
    if event_date and start_clock and end_clock and timezone_obj:
        local_date_dt = datetime.datetime.combine(event_date, datetime.time.min)
        local_start_dt = datetime.datetime.combine(event_date, start_clock)
        local_end_dt = datetime.datetime.combine(event_date, end_clock)

        date = _local_datetime_to_utc(local_date_dt, timezone_obj.offset)
        start_time = _local_datetime_to_utc(local_start_dt, timezone_obj.offset)
        end_time = _local_datetime_to_utc(local_end_dt, timezone_obj.offset)

    if request_type_id and not RequestType.objects.filter(id=request_type_id, tenant_id=tenant_id).exists():
        errors.append(
            f"request_type_id does not exist for tenant '{tenant_name}': {request_type_id}"
        )

    if event_type_id and not EventType.objects.filter(id=event_type_id, tenant_id=tenant_id).exists():
        errors.append(
            f"event_type_id does not exist for tenant '{tenant_name}': {event_type_id}"
        )

    client_id = _capture(_optional_int, row.get("client_id"), "client_id")
    if client_id and not Client.objects.filter(id=client_id, tenant_id=tenant_id).exists():
        errors.append(f"client_id does not exist for tenant '{tenant_name}': {client_id}")

    distributor_id = _capture(_optional_int, row.get("distributor_id"), "distributor_id")
    if distributor_id and not Distributor.objects.filter(id=distributor_id, tenant_id=tenant_id).exists():
        errors.append(f"distributor_id does not exist for tenant '{tenant_name}': {distributor_id}")

    retailer_id = _capture(_optional_int, row.get("retailer_id"), "retailer_id")
    if retailer_id and not Retailer.objects.filter(id=retailer_id, tenant_id=tenant_id).exists():
        errors.append(f"retailer_id does not exist for tenant '{tenant_name}': {retailer_id}")

    store_manager_id = _capture(_optional_int, row.get("store_manager_id"), "store_manager_id")
    if store_manager_id and not RequestStoreManager.objects.filter(id=store_manager_id).exists():
        errors.append(f"store_manager_id does not exist: {store_manager_id}")

    product_ids = _capture(_parse_int_list, row.get("product_ids")) or []
    if product_ids:
        unique_product_ids = sorted(set(product_ids))
        existing_product_ids = set(
            Product.objects.filter(id__in=unique_product_ids, tenant_id=tenant_id).values_list("id", flat=True)
        )
        missing_product_ids = [pid for pid in unique_product_ids if pid not in existing_product_ids]
        if missing_product_ids:
            errors.append(
                f"product_ids not found for tenant '{tenant_name}': "
                + ", ".join(str(pid) for pid in missing_product_ids)
            )

    table_size = _capture(_optional_int, row.get("table_size"), "table_size")
    create_detail = table_size is not None
    is_table_needed = table_size is not None
    if table_size is not None and table_size <= 0:
        errors.append("table_size must be greater than 0.")

    city_name = _optional_str(row.get("city"))
    if city_name is None:
        city_name = _optional_str(row.get("city_code"))

    location_id: int | None = None
    if city_name:
        location_qs = Location.objects.filter(tenant_id=tenant_id, name__iexact=city_name).order_by("id")
        location_count = location_qs.count()
        if location_count == 0:
            errors.append(f"city does not exist for tenant '{tenant_name}': {city_name}")
        elif location_count > 1:
            errors.append("Multiple cities found with that name. Please use a unique city name.")
        else:
            location_id = location_qs.first().id
    else:
        location_id = _capture(_optional_int, row.get("location_id"), "location_id")

    if location_id:
        location = Location.objects.filter(id=location_id, tenant_id=tenant_id).first()
        if not location:
            errors.append(f"location_id does not exist for tenant '{tenant_name}': {location_id}")
            state_id = None
        else:
            state_id = location.state_id
    else:
        state_id = None

    distributor_name = _optional_str(row.get("distributor_name"))
    retailer_name = _optional_str(row.get("retailer_name"))

    if distributor_name and not distributor_id:
        if not location_id:
            errors.append("city is required to match distributor_name.")
        else:
            distributor_qs = Distributor.objects.filter(
                tenant_id=tenant_id,
                location_id=location_id,
                location__state_id=state_id,
            ).order_by("id")
            try:
                matched_distributor = _find_best_name_match(
                    queryset=distributor_qs,
                    provided_name=distributor_name,
                    entity_label="distributor",
                )
            except ValueError as exc:
                errors.append(str(exc))
                matched_distributor = None
            if not matched_distributor:
                errors.append("No distributor match found for distributor_name + city + state.")
            else:
                distributor_id = matched_distributor.id

    if retailer_name and not retailer_id:
        if not location_id:
            errors.append("city is required to match retailer_name.")
        else:
            retailer_qs = Retailer.objects.filter(
                tenant_id=tenant_id,
                location_id=location_id,
                location__state_id=state_id,
            ).order_by("id")
            try:
                matched_retailer = _find_best_name_match(
                    queryset=retailer_qs,
                    provided_name=retailer_name,
                    entity_label="retailer",
                )
            except ValueError as exc:
                errors.append(str(exc))
                matched_retailer = None
            if not matched_retailer:
                errors.append("No retailer match found for retailer_name + city + state.")
            else:
                retailer_id = matched_retailer.id

    if errors:
        raise ValueError(" | ".join(errors))

    coordinates_request = [latitude, longitude] if latitude is not None and longitude is not None else []
    coordinates_event = [latitude, longitude] if latitude is not None and longitude is not None else None

    return {
        "name": name,
        "date": date,
        "start_time": start_time,
        "end_time": end_time,
        "address": address,
        "coordinates_request": coordinates_request,
        "coordinates_event": coordinates_event,
        "notes": _optional_str(row.get("notes")),
        "requestor_email": _optional_str(row.get("requestor_email")),
        "client_name": _optional_str(row.get("client_name")),
        "client_email": _optional_str(row.get("client_email")),
        "distributor_name": distributor_name,
        "distributor_email": _optional_str(row.get("distributor_email")),
        "retailer_name": retailer_name,
        "retailer_address": _optional_str(row.get("retailer_address")),
        "retailer_store_contact": _optional_str(row.get("retailer_store_contact")),
        "location_id": location_id,
        "store_manager_name": _optional_str(row.get("store_manager_name")),
        "store_manager_phone": _optional_str(row.get("store_manager_phone")),
        "timezone_id": timezone_id,
        "request_type_id": request_type_id,
        "event_type_id": event_type_id,
        "client_id": client_id,
        "distributor_id": distributor_id,
        "retailer_id": retailer_id,
        "store_manager_id": store_manager_id,
        "product_ids": product_ids,
        "is_table_needed": is_table_needed,
        "table_size": table_size,
        "create_detail": create_detail,
    }


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _required_str(row: dict[str, Any], field_name: str) -> str:
    value = row.get(field_name)
    if _is_empty(value):
        raise ValueError(f"{field_name} is required.")
    return str(value).strip()


def _optional_str(value: Any) -> str | None:
    if _is_empty(value):
        return None
    return str(value).strip()


def _required_date(row: dict[str, Any], field_name: str) -> datetime.date:
    value = row.get(field_name)
    if _is_empty(value):
        raise ValueError(f"{field_name} is required.")

    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"{field_name} must use date format mm/dd/yyyy.")


def _required_time(row: dict[str, Any], field_name: str) -> datetime.time:
    value = row.get(field_name)
    if _is_empty(value):
        raise ValueError(f"{field_name} is required.")

    if isinstance(value, datetime.datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, datetime.time):
        return value.replace(second=0, microsecond=0)

    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        seconds_in_day = int(float(value) * 24 * 60 * 60)
        hours = seconds_in_day // 3600
        minutes = (seconds_in_day % 3600) // 60
        return datetime.time(hour=hours, minute=minutes)

    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.datetime.strptime(text, fmt).time().replace(second=0, microsecond=0)
        except ValueError:
            continue

    raise ValueError(f"{field_name} must use time format HH:MM (example 14:00).")


def _optional_float(value: Any, field_name: str) -> float | None:
    if _is_empty(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field_name} is not a valid number.")


def _optional_int(value: Any, field_name: str) -> int | None:
    if _is_empty(value):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field_name} is not a valid integer.")


def _parse_int_list(value: Any) -> list[int]:
    if _is_empty(value):
        return []
    if isinstance(value, (list, tuple)):
        items = value
    else:
        normalized = str(value).replace("|", ",").replace(";", ",")
        items = [part.strip() for part in normalized.split(",")]
    parsed: list[int] = []
    for item in items:
        if _is_empty(item):
            continue
        try:
            parsed.append(int(item))
        except (TypeError, ValueError):
            raise ValueError("product_ids must be a comma-separated list of integers.")
    return parsed


def _normalize_offset_minutes(offset_value: int | None) -> int:
    if offset_value is None:
        return 0
    value = int(offset_value)
    if abs(value) > 24:
        return value
    return value * 60


def _local_datetime_to_utc(
    local_dt: datetime.datetime,
    timezone_offset_value: int | None,
) -> datetime.datetime:
    offset_minutes = _normalize_offset_minutes(timezone_offset_value)
    utc_dt = local_dt - datetime.timedelta(minutes=offset_minutes)
    return utc_dt.replace(tzinfo=datetime.timezone.utc)


def _normalize_name(value: str | None) -> str:
    return (value or "").strip().lower()


def _name_similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, _normalize_name(a), _normalize_name(b)).ratio()


def _find_best_name_match(queryset, provided_name: str, entity_label: str):
    provided = _normalize_name(provided_name)
    if not provided:
        return None

    candidates = list(queryset)
    if not candidates:
        return None

    like_candidates = []
    for candidate in candidates:
        db_name = _normalize_name(getattr(candidate, "name", ""))
        if not db_name:
            continue
        if provided in db_name or db_name in provided:
            like_candidates.append(candidate)

    pool = like_candidates if like_candidates else candidates
    scored = [
        (candidate, _name_similarity(provided_name, getattr(candidate, "name", "")))
        for candidate in pool
    ]
    scored.sort(key=lambda item: item[1], reverse=True)

    best_candidate, best_score = scored[0]
    if best_score < 0.5:
        return None

    if len(scored) > 1 and abs(best_score - scored[1][1]) < 0.03:
        raise ValueError(
            f"Multiple {entity_label} matches with similar names. Please provide a more specific name."
        )

    return best_candidate
