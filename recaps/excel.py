from __future__ import annotations

from io import BytesIO
from typing import Iterable
from openpyxl import Workbook

from utils.gcs import extract_blob_name_from_url

INVALID_SHEET_CHARS = set(r'[]:*?/\\')

def _format_dt(value) -> str:
    if not value:
        return ""
    try:
        return value.isoformat()
    except Exception:
        return str(value)


def _format_date_mdy(value) -> str:
    if not value:
        return ""
    try:
        return value.strftime("%m/%d/%Y")
    except Exception:
        return str(value)


def _safe(value) -> str:
    if value is None:
        return ""
    return str(value)


def _format_recap_status(approved: bool) -> str:
    return "Approved" if approved else "Pending"


def _format_user_name(user) -> str:
    if not user:
        return ""
    first_name = getattr(user, "first_name", "") or ""
    last_name = getattr(user, "last_name", "") or ""
    full_name = f"{first_name} {last_name}".strip()
    if full_name:
        return full_name
    return getattr(user, "username", None) or str(user)


def _get_retailer_name(recap) -> str:
    retailer = getattr(recap, "retailer", None)
    if retailer and getattr(retailer, "name", None):
        return retailer.name
    event = getattr(recap, "event", None)
    request = getattr(event, "request", None) if event else None
    request_retailer = getattr(request, "retailer", None) if request else None
    return getattr(request_retailer, "name", "") or ""


def _get_distributor_name(recap) -> str:
    event = getattr(recap, "event", None)
    request = getattr(event, "request", None) if event else None
    distributor = getattr(request, "distributor", None) if request else None
    return getattr(distributor, "name", "") or ""


def _get_retailer_state_name(recap) -> str:
    retailer = getattr(recap, "retailer", None)
    if not retailer:
        event = getattr(recap, "event", None)
        request = getattr(event, "request", None) if event else None
        retailer = getattr(request, "retailer", None) if request else None
    location = getattr(retailer, "location", None) if retailer else None
    state = getattr(location, "state", None) if location else None
    if state and getattr(state, "name", None):
        return state.name

    event = getattr(recap, "event", None)
    request = getattr(event, "request", None) if event else None
    distributor = getattr(event, "distributor", None) if event else None
    if not distributor:
        distributor = getattr(request, "distributor", None) if request else None
    distributor_location = getattr(distributor, "location", None) if distributor else None
    distributor_state = (
        getattr(distributor_location, "state", None)
        if distributor_location
        else getattr(distributor, "state", None)
    )
    return getattr(distributor_state, "name", "") or ""


def _related_items(instance, primary_attr: str, fallback_attr: str | None = None) -> list:
    relation = getattr(instance, primary_attr, None)
    if relation is None and fallback_attr:
        relation = getattr(instance, fallback_attr, None)
    if relation is None:
        return []
    try:
        return list(relation.all())
    except Exception:
        return []


def _file_field_value(recap_file):
    file_value = getattr(recap_file, "file", None)
    if file_value:
        return file_value
    return getattr(recap_file, "url", None)


def _sanitize_sheet_title(title: str, used_titles: set[str]) -> str:
    cleaned = "".join("_" if char in INVALID_SHEET_CHARS else char for char in title)
    cleaned = (cleaned or "Section").strip() or "Section"
    cleaned = cleaned[:31]
    if cleaned not in used_titles:
        used_titles.add(cleaned)
        return cleaned

    base = cleaned[:28] or "Sec"
    counter = 2
    candidate = f"{base}_{counter}"
    while candidate in used_titles:
        counter += 1
        candidate = f"{base}_{counter}"
    used_titles.add(candidate)
    return candidate


def build_recaps_xlsx(
    recaps: Iterable[object],
    frontend_base_url: str | None = None,
) -> bytes:
    """Build an Excel report containing recap data and related tables."""
    recaps_list = list(recaps)
    wb = Workbook()
    has_standard_recaps = any(
        hasattr(recap, "consumer_engagements") for recap in recaps_list
    )

    recap_sheet = wb.active
    recap_sheet.title = "Recaps"
    recap_sheet.append(
        [
            "recap_uuid",
            "recap_name",
            "status",
            "event_uuid",
            "event_name",
            "event_date",
            "event_address",
            "retailer",
            "state",
            "distributor",
            "ambassador",
            "job",
            "total_engagements",
            "products_sold",
            "total_cans_sold",
            "total_packs_sold",
            "total_earnings",
            "account_spend_amount",
            "created_at",
            "updated_at",
        ]
    )

    engagements_sheet = None
    if has_standard_recaps:
        engagements_sheet = wb.create_sheet(title="ConsumerEngagements")
        engagements_sheet.append(
            [
                "recap_uuid",
                "recap_name",
                "total_consumer",
                "first_time_consumers",
                "brand_aware_consumers",
                "willing_to_purchase_consumers",
                "not_willing_consumers",
            ]
        )

    samples_sheet = wb.create_sheet(title="ProductSamples")
    samples_sheet.append(
        [
            "recap_uuid",
            "recap_name",
            "product",
            "quantity",
        ]
    )

    sales_sheet = wb.create_sheet(title="SalesPerformance")
    sales_sheet.append(
        [
            "recap_uuid",
            "recap_name",
            "product",
            "type_of_good",
            "price",
        ]
    )

    consumer_feedback_sheet = None
    if has_standard_recaps:
        consumer_feedback_sheet = wb.create_sheet(title="ConsumerFeedback")
        consumer_feedback_sheet.append(
            [
                "recap_uuid",
                "recap_name",
                "demographics",
                "feedback",
                "quotes",
                "positive_stories",
                "reasons_to_decline",
            ]
        )

    account_feedback_sheet = None
    if has_standard_recaps:
        account_feedback_sheet = wb.create_sheet(title="AccountFeedback")
        account_feedback_sheet.append(
            [
                "recap_uuid",
                "recap_name",
                "do_differently_feedback",
                "feedback",
                "corpo_card",
            ]
        )

    files_sheet = wb.create_sheet(title="RecapFiles")
    files_sheet.append(
        [
            "recap_uuid",
            "recap_name",
            "file_name",
            "file_type",
            "file_category",
            "approved",
            "file_path",
        ]
    )

    section_definitions: dict[object, dict[str, object]] = {}
    used_sheet_titles = {
        "Recaps",
        "ProductSamples",
        "SalesPerformance",
        "RecapFiles",
    }
    if has_standard_recaps:
        used_sheet_titles.update(
            {"ConsumerEngagements", "ConsumerFeedback", "AccountFeedback"}
        )
    for recap in recaps_list:
        template_fields = _related_items(
            getattr(recap, "custom_recap_template", None),
            "custom_field",
        )
        value_fields = [
            getattr(item, "custom_field", None)
            for item in _related_items(recap, "custom_field_value")
            if getattr(item, "custom_field", None) is not None
        ]
        for custom_field in [*template_fields, *value_fields]:
            section = getattr(custom_field, "recap_section", None)
            section_id = getattr(section, "id", None)
            if section_id is None:
                continue
            section_entry = section_definitions.setdefault(
                section_id,
                {
                    "section_name": getattr(section, "name", None) or "Section",
                    "fields": {},
                },
            )
            section_entry["fields"][getattr(custom_field, "id", None)] = custom_field

    section_sheets: dict[object, tuple[object, list[object]]] = {}
    for section_id, section_entry in section_definitions.items():
        sheet_title = _sanitize_sheet_title(
            str(section_entry["section_name"]),
            used_sheet_titles,
        )
        section_sheet = wb.create_sheet(title=sheet_title)
        section_fields = sorted(
            section_entry["fields"].values(),
            key=lambda field: (
                getattr(field, "id", 0) is None,
                getattr(field, "id", 0) or 0,
            ),
        )
        section_sheet.append(
            [
                "recap_uuid",
                "recap_name",
                *[_safe(getattr(field, "name", None)) for field in section_fields],
            ]
        )
        section_sheets[section_id] = (section_sheet, section_fields)

    for recap in recaps_list:
        ambassador_user = None
        if getattr(recap, "ambassador", None) and getattr(recap.ambassador, "user", None):
            ambassador_user = recap.ambassador.user
        recap_approved = bool(getattr(recap, "approved", False))

        recap_sheet.append(
            [
                _safe(getattr(recap, "uuid", None)),
                _safe(getattr(recap, "name", None)),
                _format_recap_status(recap_approved),
                _safe(getattr(getattr(recap, "event", None), "uuid", None)),
                _safe(getattr(getattr(recap, "event", None), "name", None)),
                _format_date_mdy(getattr(getattr(recap, "event", None), "date", None)),
                _safe(getattr(getattr(recap, "event", None), "address", None)),
                _get_retailer_name(recap),
                _get_retailer_state_name(recap),
                _get_distributor_name(recap),
                _format_user_name(ambassador_user),
                _safe(getattr(getattr(recap, "job", None), "name", None)),
                _safe(getattr(recap, "total_engagements", None)),
                _safe(getattr(recap, "products_sold", None)),
                _safe(getattr(recap, "total_cans_sold", None)),
                _safe(getattr(recap, "total_packs_sold", None)),
                _safe(getattr(recap, "total_earnings", None)),
                _safe(getattr(recap, "account_spend_amount", None)),
                _format_dt(getattr(recap, "created_at", None)),
                _format_dt(getattr(recap, "updated_at", None)),
            ]
        )

        if engagements_sheet is not None:
            for engagement in _related_items(recap, "consumer_engagements"):
                engagements_sheet.append(
                    [
                        _safe(getattr(recap, "uuid", None)),
                        _safe(getattr(recap, "name", None)),
                        _safe(getattr(engagement, "total_consumer", None)),
                        _safe(getattr(engagement, "first_time_consumers", None)),
                        _safe(getattr(engagement, "brand_aware_consumers", None)),
                        _safe(getattr(engagement, "willing_to_purchase_consumers", None)),
                        _safe(getattr(engagement, "not_willing_consumers", None)),
                    ]
                )

        for sample in _related_items(
            recap, "product_samples", "custom_recap_product_sample"
        ):
            samples_sheet.append(
                [
                    _safe(getattr(recap, "uuid", None)),
                    _safe(getattr(recap, "name", None)),
                    _safe(getattr(getattr(sample, "product", None), "name", None)),
                    _safe(getattr(sample, "quantity", None)),
                ]
            )

        for sale in _related_items(
            recap, "sales_performance", "custom_recap_sale_performance"
        ):
            sales_sheet.append(
                [
                    _safe(getattr(recap, "uuid", None)),
                    _safe(getattr(recap, "name", None)),
                    _safe(getattr(getattr(sale, "product", None), "name", None)),
                    _safe(getattr(getattr(sale, "type_of_good", None), "name", None)),
                    _safe(getattr(sale, "price", None)),
                ]
            )

        if consumer_feedback_sheet is not None:
            for feedback in _related_items(recap, "consumer_feedback"):
                consumer_feedback_sheet.append(
                    [
                        _safe(getattr(recap, "uuid", None)),
                        _safe(getattr(recap, "name", None)),
                        _safe(getattr(feedback, "demographics", None)),
                        _safe(getattr(feedback, "feedback", None)),
                        _safe(getattr(feedback, "quotes", None)),
                        _safe(getattr(feedback, "positive_stories", None)),
                        _safe(getattr(feedback, "reasons_to_decline", None)),
                    ]
                )

        if account_feedback_sheet is not None:
            for feedback in _related_items(recap, "account_feedback"):
                account_feedback_sheet.append(
                    [
                        _safe(getattr(recap, "uuid", None)),
                        _safe(getattr(recap, "name", None)),
                        _safe(getattr(feedback, "do_differently_feedback", None)),
                        _safe(getattr(feedback, "feedback", None)),
                        _safe(getattr(feedback, "corpo_card", None)),
                    ]
                )

        for recap_file in _related_items(recap, "recap_files", "custom_recap_files"):
            file_path = _file_field_value(recap_file)
            blob_name = extract_blob_name_from_url(str(file_path)) if file_path else None
            signed_url = ""
            if frontend_base_url:
                recap_file_id = getattr(recap_file, "uuid", None) or getattr(
                    recap_file, "id", None
                )
                if recap_file_id is not None:
                    signed_url = f"{frontend_base_url}/recap/file/{recap_file_id}"
            display_name = getattr(recap_file, "name", None) or ""
            if not display_name and file_path:
                try:
                    display_name = str(file_path).rsplit("/", 1)[-1]
                except Exception:
                    display_name = ""
            files_sheet.append(
                [
                    _safe(getattr(recap, "uuid", None)),
                    _safe(getattr(recap, "name", None)),
                    _safe(getattr(recap_file, "name", None)),
                    _safe(getattr(getattr(recap_file, "file_type", None), "name", None)),
                    _safe(
                        getattr(
                            getattr(recap_file, "file_recap_category", None),
                            "name",
                            None,
                        )
                    ),
                    bool(getattr(recap_file, "approved", False)),
                    display_name or signed_url,
                ]
            )
            if signed_url:
                cell = files_sheet.cell(row=files_sheet.max_row, column=7)
                cell.hyperlink = signed_url
                cell.style = "Hyperlink"

        if section_sheets:
            value_by_field_id = {
                getattr(getattr(custom_field_value, "custom_field", None), "id", None): _safe(
                    getattr(custom_field_value, "value", None)
                )
                for custom_field_value in _related_items(recap, "custom_field_value")
            }
            for section_id, (section_sheet, section_fields) in section_sheets.items():
                row = [
                    _safe(getattr(recap, "uuid", None)),
                    _safe(getattr(recap, "name", None)),
                ]
                has_section_field = False
                for field in section_fields:
                    field_id = getattr(field, "id", None)
                    field_section = getattr(field, "recap_section", None)
                    if getattr(field_section, "id", None) != section_id:
                        row.append("")
                        continue
                    has_section_field = True
                    row.append(value_by_field_id.get(field_id, ""))
                if has_section_field:
                    section_sheet.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
