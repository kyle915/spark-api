"""Coverage for the tenant-wide FIELD-LEVEL recap export.

``build_recap_field_export(tenant)`` flattens a tenant's recaps into one row per
recap: identity columns, then one column per template field in section order,
then one column per ``FileRecapCategory`` holding that category's public URLs.
This pins the behaviours that make it trustworthy as a client deliverable:

  * column order follows template → section.order → field.order, image-type
    fields excluded (they hold a blob path, not an answer),
  * a field name shared by two templates collapses to ONE column, because
    values are keyed by normalized name and a second column would silently
    duplicate the first,
  * a multiselect answer (JSON array) renders as a readable comma list,
  * unanswered fields stay BLANK — never 0, which would read as measured,
  * files group under their real category and resolve to public GCS URLs built
    via ``public_url``/``extract_blob_name_from_url`` (not hand-concatenated),
  * receipt-looking files sitting in a NON-receipt category are flagged — Girl
    Beer's receipts once filed themselves under a photo category,
  * the date window scopes by EVENT date, not created/imported date,
  * events with no recap are COUNTED, not emitted as blank rows.

The XLSX/CSV rendering (:mod:`utils.workbook`) is exercised separately at the
bottom — it's pure and Django-free, so it's cheap to assert on directly.
"""
from __future__ import annotations

from datetime import timedelta

import pytest
from django.utils import timezone

from ambassadors.models import FileType
from ambassadors.tests.base import AmbassadorsGraphQLTestCase
from recaps import models as recap_models
from recaps.recap_field_export import UNCATEGORIZED_LABEL, build_recap_field_export


@pytest.mark.django_db(transaction=True)
class TestRecapFieldExport(AmbassadorsGraphQLTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, db, settings):
        # `public_url` returns None with no bucket configured, which would drop
        # every file from the export — so the file assertions below need a
        # bucket set. Production reads this from the environment.
        settings.GS_BUCKET_NAME = "sparkio-test"
        self.system_user = self.get_system_user()
        self.now = timezone.now()
        self.tenant = self.create_tenant(name="GB Field Export")
        self.event_type = self.create_event_type(name="Retail Sampling", tenant=self.tenant)

        self.number_type = recap_models.CustomRecapFieldType.objects.create(
            name="number", created_by=self.system_user
        )
        self.multiselect_type = recap_models.CustomRecapFieldType.objects.create(
            name="multiselect", created_by=self.system_user
        )
        self.text_type = recap_models.CustomRecapFieldType.objects.create(
            name="text", created_by=self.system_user
        )
        self.image_type = recap_models.CustomRecapFieldType.objects.create(
            name="image", created_by=self.system_user
        )
        self.file_type = FileType.objects.create(name="image", created_by=self.system_user)

        self.template = recap_models.CustomRecapTemplate.objects.create(
            name="GB Recap",
            event_type=self.event_type,
            tenant=self.tenant,
            created_by=self.system_user,
        )
        self.sales = recap_models.RecapSection.objects.create(
            name="Sales Figures", tenant=self.tenant, created_by=self.system_user, order=1
        )
        self.demo = recap_models.RecapSection.objects.create(
            name="Demographics", tenant=self.tenant, created_by=self.system_user, order=2
        )

        self.samples = self._field("Total Samples Given Out", self.number_type, self.sales, 1)
        # Image field in the middle of the ordering — must be EXCLUDED.
        self._field("Table setup pictures", self.image_type, self.sales, 2)
        self.consumers = self._field("Consumers Engaged", self.number_type, self.sales, 3)
        self.market = self._field("Products Sampled", self.multiselect_type, self.demo, 1)
        self.notes = self._field("Anything else?", self.text_type, self.demo, 2)

        # Tenant file categories, seeded in the production default order.
        self.cats = {}
        for name in ("Sampling photos", "Table setup", "Receipts"):
            self.cats[name] = recap_models.FileRecapCategory.objects.create(
                name=name, tenant=self.tenant, created_by=self.system_user
            )

    # ── builders ───────────────────────────────────────────────────────
    def _field(self, name, field_type, section, order, template=None):
        return recap_models.CustomField.objects.create(
            name=name,
            custom_recap_template=template or self.template,
            custom_field_type=field_type,
            recap_section=section,
            created_by=self.system_user,
            order=order,
        )

    def _recap(self, idx, values=(), *, days=0, approved=True, event=None):
        event = event or self.create_event(
            name=f"Whole Foods Store {idx:03d}",
            tenant=self.tenant,
            date=self.now + timedelta(days=days),
        )
        recap = recap_models.CustomRecap.objects.create(
            name=f"recap {idx:03d}",
            approved=approved,
            event=event,
            tenant=self.tenant,
            custom_recap_template=self.template,
            submitted_at=self.now + timedelta(days=days),
            created_by=self.system_user,
            updated_by=self.system_user,
        )
        for field, value in values:
            recap_models.CustomFieldValue.objects.create(
                value=value,
                custom_recap=recap,
                custom_field=field,
                created_by=self.system_user,
            )
        return recap

    def _file(self, recap, blob, category_name):
        return recap_models.CustomRecapFile.objects.create(
            name=blob.rsplit("/", 1)[-1],
            url=blob,
            custom_recap=recap,
            file_type=self.file_type,
            file_recap_category=self.cats[category_name] if category_name else None,
            created_by=self.system_user,
        )

    @staticmethod
    def _cols(payload):
        return [c["header"] for c in payload["columns"]]

    @staticmethod
    def _row_dict(payload, index=0):
        headers = [c["header"] for c in payload["columns"]]
        return dict(zip(headers, payload["rows"][index]))

    # ── columns ────────────────────────────────────────────────────────
    def test_columns_ordered_by_section_and_exclude_image_fields(self):
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        headers = self._cols(payload)

        # Identity block leads, in a stable order.
        assert headers[:4] == ["Recap ID", "Recap #", "Request ID", "Event"]
        assert "Approval Status" in headers

        # Field columns follow section.order then field.order; image dropped.
        field_headers = [
            c["header"] for c in payload["columns"] if c["key"].startswith("field::")
        ]
        assert field_headers == [
            "Total Samples Given Out",
            "Consumers Engaged",
            "Products Sampled",
            "Anything else?",
        ]
        assert "Table setup pictures" not in headers

        # File columns: one per category in tenant creation order, then total.
        assert headers[-4:] == ["Sampling photos", "Table setup", "Receipts", "Total Files"]

    def test_field_columns_carry_their_section_as_group(self):
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        by_header = {c["header"]: c for c in payload["columns"]}
        assert by_header["Total Samples Given Out"]["group"] == "Sales Figures"
        assert by_header["Products Sampled"]["group"] == "Demographics"
        assert by_header["Sampling photos"]["group"] == "Files"
        assert by_header["Event"]["group"] == "Identity"

    def test_duplicate_field_name_across_templates_collapses_to_one_column(self):
        other = recap_models.CustomRecapTemplate.objects.create(
            name="GB Recap v2",
            event_type=self.event_type,
            tenant=self.tenant,
            created_by=self.system_user,
        )
        # Same name, different template — one column, or the second silently
        # duplicates the first (values are keyed by normalized name).
        self._field("Total Samples Given Out", self.number_type, self.sales, 1, template=other)
        self._field("Brand New Question", self.text_type, self.demo, 9, template=other)

        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        headers = self._cols(payload)
        assert headers.count("Total Samples Given Out") == 1
        assert "Brand New Question" in headers
        assert payload["diagnostics"]["duplicate_field_names_collapsed"] == [
            "Total Samples Given Out"
        ]

    # ── row values ─────────────────────────────────────────────────────
    def test_row_values_align_and_multiselect_renders_as_comma_list(self):
        self._recap(
            1,
            [
                (self.samples, "120"),
                (self.consumers, "310"),
                (self.market, '["Hazy IPA", "Lager"]'),
            ],
        )
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        assert payload["meta"]["row_count"] == 1
        assert len(payload["rows"][0]) == len(payload["columns"])

        col = self._row_dict(payload)
        assert col["Total Samples Given Out"] == "120"
        assert col["Consumers Engaged"] == "310"
        # Multiselect JSON → readable list, not raw JSON.
        assert col["Products Sampled"] == "Hazy IPA, Lager"
        assert col["Approval Status"] == "Approved"
        assert col["Event"] == "Whole Foods Store 001"

    def test_unanswered_field_is_blank_never_zero(self):
        # The landmine: a zero reads as a measured result. An unanswered
        # question must be empty.
        self._recap(2, [(self.samples, "99")])
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        col = self._row_dict(payload)
        assert col["Total Samples Given Out"] == "99"
        assert col["Consumers Engaged"] == ""
        assert col["Products Sampled"] == ""
        assert col["Anything else?"] == ""
        assert 0 not in (col["Consumers Engaged"], col["Products Sampled"])

    def test_grain_is_one_row_per_recap(self):
        self._recap(1)
        self._recap(2)
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        assert payload["grain"] == "recap"
        assert payload["meta"]["row_count"] == 2

    def test_events_without_recap_are_counted_not_rowed(self):
        self._recap(1)
        # Two events that never got a recap — the coverage gap.
        self.create_event(name="No recap A", tenant=self.tenant, date=self.now)
        self.create_event(name="No recap B", tenant=self.tenant, date=self.now)
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        assert payload["meta"]["row_count"] == 1
        assert payload["diagnostics"]["events_without_recap"] == 2

    # ── files ──────────────────────────────────────────────────────────
    def test_files_group_under_their_category_as_public_urls(self):
        recap = self._recap(1, [(self.samples, "10")])
        self._file(recap, "recaps/abc/photo-one.jpg", "Sampling photos")
        self._file(recap, "recaps/abc/photo-two.jpg", "Sampling photos")
        self._file(recap, "recaps/receipts/abc/receipt-1.jpg", "Receipts")

        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        col = self._row_dict(payload)

        assert len(col["Sampling photos"]) == 2
        assert len(col["Receipts"]) == 1
        assert col["Table setup"] == []
        assert col["Total Files"] == 3

        # Built through public_url — a real absolute GCS object URL, not a
        # bare blob name and not a signed URL with a query string.
        for url in col["Sampling photos"] + col["Receipts"]:
            assert url.startswith("https://storage.googleapis.com/")
            assert "?" not in url, "public-URL mode: links must not be signed/expiring"
        assert col["Sampling photos"][0].endswith("recaps/abc/photo-one.jpg")

        # Flat per-file list backs the clickable detail tab.
        assert payload["meta"]["file_count"] == 3
        cats = sorted(f["category"] for f in payload["files"])
        assert cats == ["Receipts", "Sampling photos", "Sampling photos"]

    def test_receipt_looking_file_in_photo_category_is_flagged(self):
        # Girl Beer's actual historical bug: receipts filed under a photo
        # category. Grouping without checking puts receipts in the photo column.
        recap = self._recap(1)
        self._file(recap, "recaps/abc/IMG_0042.jpg", "Sampling photos")
        self._file(recap, "recaps/abc/receipt-albertsons.jpg", "Sampling photos")

        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        flagged = payload["diagnostics"]["receipt_looking_files_outside_receipt_categories"]
        assert flagged["count"] == 1
        assert flagged["samples"][0]["category"] == "Sampling photos"
        assert "receipt" in flagged["samples"][0]["name"].lower()

    def test_clean_categories_produce_no_warning(self):
        recap = self._recap(1)
        self._file(recap, "recaps/abc/IMG_0042.jpg", "Sampling photos")
        self._file(recap, "recaps/receipts/abc/receipt-1.jpg", "Receipts")
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        assert (
            payload["diagnostics"]["receipt_looking_files_outside_receipt_categories"]["count"]
            == 0
        )

    def test_uncategorized_file_gets_its_own_column(self):
        recap = self._recap(1)
        self._file(recap, "recaps/abc/orphan.jpg", None)
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        headers = self._cols(payload)
        assert UNCATEGORIZED_LABEL in headers
        col = self._row_dict(payload)
        assert len(col[UNCATEGORIZED_LABEL]) == 1

    def test_unlinkable_files_are_counted_not_silently_dropped(self, settings):
        # With no bucket configured `public_url` returns None. The export must
        # SAY it couldn't link the files rather than quietly ship a deliverable
        # with the photos missing.
        settings.GS_BUCKET_NAME = ""
        recap = self._recap(1)
        self._file(recap, "recaps/abc/photo.jpg", "Sampling photos")
        self._file(recap, "recaps/abc/other.jpg", "Sampling photos")

        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        assert payload["meta"]["file_count"] == 0
        assert payload["diagnostics"]["files_without_a_link"] == 2

    def test_linkable_files_report_no_missing_links(self):
        recap = self._recap(1)
        self._file(recap, "recaps/abc/photo.jpg", "Sampling photos")
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        assert payload["diagnostics"]["files_without_a_link"] == 0

    def test_no_uncategorized_column_when_every_file_is_filed(self):
        recap = self._recap(1)
        self._file(recap, "recaps/abc/photo.jpg", "Sampling photos")
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        assert UNCATEGORIZED_LABEL not in self._cols(payload)

    # ── date window ────────────────────────────────────────────────────
    def test_window_scopes_by_event_date(self):
        # Events 10 days apart; window catches only the middle one.
        self._recap(1, [(self.samples, "1")], days=-20)
        self._recap(2, [(self.samples, "2")], days=0)
        self._recap(3, [(self.samples, "3")], days=20)

        today = self.now.date()
        payload = build_recap_field_export(
            self.tenant,
            start=(today - timedelta(days=5)).isoformat(),
            end=(today + timedelta(days=5)).isoformat(),
            resolve_heic=False,
        )
        assert payload["meta"]["row_count"] == 1
        assert self._row_dict(payload)["Total Samples Given Out"] == "2"
        assert payload["diagnostics"]["recaps_excluded_out_of_window"] == 2
        assert payload["window"]["scoped_by"] == "event date"

    def test_recap_with_no_event_date_is_excluded_from_a_window_and_reported(self):
        event = self.create_event(name="Undated", tenant=self.tenant, date=None)
        self._recap(9, [(self.samples, "7")], event=event)
        today = self.now.date()
        payload = build_recap_field_export(
            self.tenant,
            start=(today - timedelta(days=1)).isoformat(),
            end=(today + timedelta(days=1)).isoformat(),
            resolve_heic=False,
        )
        assert payload["meta"]["row_count"] == 0
        assert payload["diagnostics"]["recaps_without_event_date"] == 1

    def test_undated_recap_is_included_when_no_window_given(self):
        event = self.create_event(name="Undated", tenant=self.tenant, date=None)
        self._recap(9, [(self.samples, "7")], event=event)
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        assert payload["meta"]["row_count"] == 1
        assert payload["diagnostics"]["recaps_without_event_date"] == 1

    def test_bad_date_raises_value_error(self):
        with pytest.raises(ValueError):
            build_recap_field_export(self.tenant, start="07/01/2026", resolve_heic=False)

    # ── tenant isolation ───────────────────────────────────────────────
    def test_other_tenants_recaps_and_categories_are_excluded(self):
        other = self.create_tenant(name="Someone Else")
        other_type = self.create_event_type(name="Other Sampling", tenant=other)
        other_template = recap_models.CustomRecapTemplate.objects.create(
            name="Other Recap",
            event_type=other_type,
            tenant=other,
            created_by=self.system_user,
        )
        other_section = recap_models.RecapSection.objects.create(
            name="Other Section", tenant=other, created_by=self.system_user, order=1
        )
        self._field("Secret Other Field", self.text_type, other_section, 1, template=other_template)
        recap_models.FileRecapCategory.objects.create(
            name="Other Category", tenant=other, created_by=self.system_user
        )
        other_event = self.create_event(name="Other event", tenant=other, date=self.now)
        recap_models.CustomRecap.objects.create(
            name="other recap",
            event=other_event,
            tenant=other,
            custom_recap_template=other_template,
            created_by=self.system_user,
        )
        self._recap(1)

        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        headers = self._cols(payload)
        assert "Secret Other Field" not in headers
        assert "Other Category" not in headers
        assert payload["meta"]["row_count"] == 1

    # ── diagnostics ────────────────────────────────────────────────────
    def test_entirely_empty_field_columns_are_reported(self):
        self._recap(1, [(self.samples, "5")])
        payload = build_recap_field_export(self.tenant, resolve_heic=False)
        empties = payload["diagnostics"]["field_columns_entirely_empty"]
        assert "Consumers Engaged" in empties
        assert "Total Samples Given Out" not in empties


@pytest.mark.django_db(transaction=True)
class TestExportRecapFieldsCommand(AmbassadorsGraphQLTestCase):
    """The command + the marker protocol the cron endpoint parses.

    The endpoint splits stdout on RECAPFIELDS_JSON_START/END, so a change to
    either marker or to what the command prints breaks it in prod with no local
    signal. This pins the contract from the endpoint's side.
    """

    @pytest.fixture(autouse=True)
    def setup(self, db, settings):
        settings.GS_BUCKET_NAME = "sparkio-test"
        self.system_user = self.get_system_user()
        self.now = timezone.now()
        self.tenant = self.create_tenant(name="Girl Beer Cmd")
        self.event_type = self.create_event_type(name="Sampling", tenant=self.tenant)
        text_type = recap_models.CustomRecapFieldType.objects.create(
            name="text", created_by=self.system_user
        )
        template = recap_models.CustomRecapTemplate.objects.create(
            name="GB Recap",
            event_type=self.event_type,
            tenant=self.tenant,
            created_by=self.system_user,
        )
        section = recap_models.RecapSection.objects.create(
            name="Sales", tenant=self.tenant, created_by=self.system_user, order=1
        )
        field = recap_models.CustomField.objects.create(
            name="Total Samples Given Out",
            custom_recap_template=template,
            custom_field_type=text_type,
            recap_section=section,
            created_by=self.system_user,
            order=1,
        )
        event = self.create_event(name="Whole Foods LA", tenant=self.tenant, date=self.now)
        recap = recap_models.CustomRecap.objects.create(
            name="r1",
            approved=True,
            event=event,
            tenant=self.tenant,
            custom_recap_template=template,
            created_by=self.system_user,
        )
        recap_models.CustomFieldValue.objects.create(
            value="120", custom_recap=recap, custom_field=field, created_by=self.system_user
        )

    def _run(self, **kwargs) -> str:
        import io as _io

        from django.core.management import call_command

        out = _io.StringIO()
        call_command("export_recap_fields", stdout=out, **kwargs)
        return out.getvalue()

    def test_json_payload_parses_out_of_the_markers(self):
        import json as _json

        raw = self._run(tenant=str(self.tenant.id), as_json=True)
        # Exactly how digest.cron_views.DumpRecapFieldsView extracts it.
        blob = raw.split("RECAPFIELDS_JSON_START", 1)[1].split("RECAPFIELDS_JSON_END", 1)[0]
        payload = _json.loads(blob)
        assert payload["meta"]["row_count"] == 1
        assert payload["grain"] == "recap"
        assert payload["generated_at"]
        headers = [c["header"] for c in payload["columns"]]
        assert "Total Samples Given Out" in headers

    def test_report_precedes_the_markers_so_the_log_stays_readable(self):
        raw = self._run(tenant=str(self.tenant.id), as_json=True)
        report = raw.split("RECAPFIELDS_JSON_START", 1)[0]
        assert "field-level recap export" in report
        assert "one row per RECAP" in report
        assert "scoped by EVENT date" in report

    def test_dry_run_writes_no_payload(self):
        raw = self._run(tenant=str(self.tenant.id), dry_run=True)
        assert "RECAPFIELDS_JSON_START" not in raw
        assert "Dry run" in raw
        assert "rows:    1" in raw

    def test_out_writes_xlsx_and_both_csvs(self, tmp_path):
        import os

        stem = str(tmp_path / "gb")
        raw = self._run(tenant=str(self.tenant.id), out=stem)
        for suffix in (".xlsx", ".csv", "-files.csv"):
            assert os.path.getsize(stem + suffix) > 0
        assert "wrote" in raw

    def test_resolves_tenant_by_name_not_just_id(self):
        raw = self._run(tenant="Girl Beer Cmd", dry_run=True)
        assert "Girl Beer Cmd" in raw

    def test_bad_date_is_a_clean_command_error(self):
        from django.core.management.base import CommandError

        with pytest.raises(CommandError, match="bad date"):
            self._run(tenant=str(self.tenant.id), start="07/01/2026", dry_run=True)


# ───────────────────────────────────────────────────────────────────────
# Rendering — pure, no DB.
# ───────────────────────────────────────────────────────────────────────
def _payload():
    return {
        "tenant": {"id": 10, "name": "Girl Beer", "slug": "girl-beer"},
        "grain": "recap",
        "window": {"start": "2026-07-01", "end": "2026-07-31", "scoped_by": "event date"},
        "generated_at": "2026-08-03T12:00:00+00:00",
        "columns": [
            {"key": "event_name", "header": "Event", "group": "Identity", "kind": "text"},
            {
                "key": "field::total samples",
                "header": "Total Samples",
                "group": "Sales Figures",
                "kind": "text",
            },
            {
                "key": "filecat::sampling photos",
                "header": "Sampling photos",
                "group": "Files",
                "kind": "links",
            },
            {
                "key": "filecat::receipts",
                "header": "Receipts",
                "group": "Files",
                "kind": "links",
            },
            {"key": "file_total", "header": "Total Files", "group": "Files", "kind": "number"},
        ],
        "rows": [
            [
                "Whole Foods LA",
                "120",
                [
                    "https://storage.googleapis.com/b/recaps/a.jpg",
                    "https://storage.googleapis.com/b/recaps/b.jpg",
                ],
                ["https://storage.googleapis.com/b/recaps/receipts/r.jpg"],
                3,
            ],
            ["Albertsons\x0bLaguna", "", [], [], 0],
        ],
        "files": [
            {
                "recap_uuid": "u-1",
                "event_name": "Whole Foods LA",
                "event_date": "07/12/2026",
                "ba": "Dana Lane",
                "category": "Sampling photos",
                "name": "a.jpg",
                "url": "https://storage.googleapis.com/b/recaps/a.jpg",
                "approved": True,
            }
        ],
        "meta": {
            "row_count": 2,
            "column_count": 5,
            "file_count": 1,
            "identity_column_count": 1,
            "field_column_count": 1,
            "file_column_count": 3,
            "templates": {1: "GB Recap"},
        },
        "diagnostics": {
            "events_without_recap": 20,
            "files_by_category": {"Sampling photos": 2, "Receipts": 1},
            "receipt_looking_files_outside_receipt_categories": {"count": 0, "samples": []},
            "field_columns_entirely_empty": [],
        },
    }


class TestWorkbookRendering:
    def test_xlsx_has_the_three_sheets_and_a_group_header_row(self):
        from openpyxl import load_workbook
        import io as _io

        from utils.workbook import DATA_SHEET, FILES_SHEET, NOTES_SHEET, build_xlsx

        wb = load_workbook(_io.BytesIO(build_xlsx(_payload())))
        assert wb.sheetnames == [NOTES_SHEET, DATA_SHEET, FILES_SHEET]

        ws = wb[DATA_SHEET]
        # Row 1 = column group, row 2 = header, row 3+ = data.
        assert ws.cell(row=1, column=1).value == "Identity"
        assert ws.cell(row=1, column=2).value == "Sales Figures"
        assert ws.cell(row=2, column=1).value == "Event"
        assert ws.cell(row=3, column=1).value == "Whole Foods LA"
        assert ws.freeze_panes == "A3"

    def test_group_label_is_written_once_per_run_not_repeated(self):
        from openpyxl import load_workbook
        import io as _io

        from utils.workbook import DATA_SHEET, build_xlsx

        payload = _payload()
        # Three consecutive "Files" columns → the label appears on the first
        # one only, so the band reads as one group instead of "Files Files Files".
        ws = load_workbook(_io.BytesIO(build_xlsx(payload)))[DATA_SHEET]
        groups = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert groups == ["Identity", "Sales Figures", "Files", None, None]

    def test_multi_url_cell_lists_urls_and_single_url_cell_is_hyperlinked(self):
        from openpyxl import load_workbook
        import io as _io

        from utils.workbook import DATA_SHEET, build_xlsx

        ws = load_workbook(_io.BytesIO(build_xlsx(_payload())))[DATA_SHEET]
        photos = ws.cell(row=3, column=3)
        receipts = ws.cell(row=3, column=4)

        # Two files → newline-joined text (Excel allows one link per cell).
        assert photos.value.count("\n") == 1
        assert photos.value.startswith("https://storage.googleapis.com/")
        # Exactly one file → a real clickable hyperlink.
        assert receipts.hyperlink is not None
        assert receipts.hyperlink.target.endswith("/receipts/r.jpg")

    def test_files_sheet_rows_are_clickable(self):
        from openpyxl import load_workbook
        import io as _io

        from utils.workbook import FILES_SHEET, build_xlsx

        ws = load_workbook(_io.BytesIO(build_xlsx(_payload())))[FILES_SHEET]
        assert ws.cell(row=1, column=1).value == "Category"
        assert ws.cell(row=2, column=1).value == "Sampling photos"
        assert ws.cell(row=2, column=6).hyperlink is not None

    def test_notes_sheet_states_the_grain_and_the_two_landmines(self):
        from openpyxl import load_workbook
        import io as _io

        from utils.workbook import NOTES_SHEET, build_xlsx

        ws = load_workbook(_io.BytesIO(build_xlsx(_payload())))[NOTES_SHEET]
        text = "\n".join(
            str(c.value or "") for row in ws.iter_rows() for c in row
        )
        assert "One row per RECAP" in text
        # The coverage gap is stated, not padded into the grid.
        assert "20" in text
        # Blanks-aren't-zeros and samples≠consumers must both be spelled out.
        assert "does not mean zero" in text
        assert "DIFFERENT measures" in text
        # And links are promised as non-expiring, which we verified upstream.
        assert "does not expire" in text

    def test_control_characters_are_stripped_so_openpyxl_cannot_choke(self):
        from openpyxl import load_workbook
        import io as _io

        from utils.workbook import DATA_SHEET, build_xlsx

        ws = load_workbook(_io.BytesIO(build_xlsx(_payload())))[DATA_SHEET]
        assert ws.cell(row=4, column=1).value == "AlbertsonsLaguna"

    def test_grid_csv_matches_headers_and_joins_link_cells(self):
        import csv as _csv
        import io as _io

        from utils.workbook import build_grid_csv

        rows = list(_csv.reader(_io.StringIO(build_grid_csv(_payload()))))
        assert rows[0] == [
            "Event",
            "Total Samples",
            "Sampling photos",
            "Receipts",
            "Total Files",
        ]
        assert rows[1][0] == "Whole Foods LA"
        assert rows[1][2].count("\n") == 1
        assert rows[2][2] == ""

    def test_files_csv_has_one_row_per_file(self):
        import csv as _csv
        import io as _io

        from utils.workbook import build_files_csv

        rows = list(_csv.reader(_io.StringIO(build_files_csv(_payload()))))
        assert rows[0][0] == "Category"
        assert len(rows) == 2
        assert rows[1][5] == "https://storage.googleapis.com/b/recaps/a.jpg"

    def test_write_outputs_creates_all_three_files(self, tmp_path):
        from utils.workbook import write_outputs

        written = write_outputs(_payload(), str(tmp_path / "gb-recaps.xlsx"))
        assert written["xlsx"].endswith("gb-recaps.xlsx")
        for path in written.values():
            import os

            assert os.path.getsize(path) > 0
