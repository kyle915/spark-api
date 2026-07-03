from unittest.mock import patch

import pytest
from asgiref.sync import sync_to_async
from openpyxl import load_workbook
import strawberry_django  # noqa: F401
from django.test import override_settings
from io import BytesIO

from ambassadors.models import AmbassadorEvent
from events.models import EventType, Product, ProductType
from jobs.tests.base import JobsGraphQLTestCase
from recaps import models as recap_models
from recaps.envelopes import RecapApprovedNotificationMailer
from recaps.excel import build_recaps_xlsx
from recaps.mutations import _notify_recap_ready_for_review_to_admins


@pytest.mark.django_db(transaction=True)
class TestApproveRecapNotifications(JobsGraphQLTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, db):
        from config.schema_spark import schema_spark

        self.roles = self.setup_default_roles()
        self.tenant = self.create_tenant(name="Recap Tenant")

        self.spark_user = self.create_user(
            username="spark_recap@test.com",
            email="spark_recap@test.com",
            role=self.roles["spark_admin"],
            password="testpass123",
        )
        self.create_tenanted_user(user=self.spark_user, tenant=self.tenant)

        self.rmm_user = self.create_user(
            username="rmm_recap@test.com",
            email="rmm_recap@test.com",
            first_name="Rosa",
            role=self.roles["client"],
            password="testpass123",
        )
        self.create_tenanted_user(user=self.rmm_user, tenant=self.tenant)

        self.event = self.create_event(
            name="Recap Event",
            tenant=self.tenant,
            address="123 Recap St",
            rmm_asigned=self.rmm_user,
        )
        self.job_title = self.create_job_title(name="BA", tenant=self.tenant)
        self.job = self.create_job(
            name="Recap Job",
            code="RECAP-JOB-001",
            address="456 Activation Ave",
            event=self.event,
            job_title=self.job_title,
            tenant=self.tenant,
        )

        self.ambassador_user = self.create_user(
            username="recap_amb@test.com",
            email="recap_amb@test.com",
            role=self.roles["ambassador"],
            password="testpass123",
        )
        self.ambassador = self.create_ambassador(user=self.ambassador_user)
        self.create_tenanted_user(user=self.ambassador_user, tenant=self.tenant)

        self.other_ambassador_user = self.create_user(
            username="other_recap_amb@test.com",
            email="other_recap_amb@test.com",
            role=self.roles["ambassador"],
            password="testpass123",
        )
        self.other_ambassador = self.create_ambassador(user=self.other_ambassador_user)
        self.create_tenanted_user(user=self.other_ambassador_user, tenant=self.tenant)

        system_user = self.get_system_user()
        AmbassadorEvent.objects.create(
            ambassador=self.ambassador,
            event=self.event,
            tenant=self.tenant,
            created_by=system_user,
        )
        AmbassadorEvent.objects.create(
            ambassador=self.other_ambassador,
            event=self.event,
            tenant=self.tenant,
            created_by=system_user,
        )

        self.recap = recap_models.Recap.objects.create(
            name="Post activation recap",
            approved=False,
            event=self.event,
            job=self.job,
            ambassador=self.ambassador,
            created_by=self.spark_user,
            updated_by=self.spark_user,
        )

        self.schema = schema_spark
        self.endpoint_path = "/api/v1/graphql/spark"

    @pytest.mark.asyncio
    async def test_approve_recap_sends_notification_email(self):
        mutation = """
        mutation ApproveRecap($input: ApproveRecapInput!) {
            approveRecap(input: $input) {
                success
                message
                recap {
                    id
                    approved
                }
            }
        }
        """
        variables = {
            "input": {
                "id": str(self.recap.id),
                "approved": True,
            }
        }

        with patch("recaps.mutations.RecapApprovedNotificationMailer.send") as mock_send:
            result = await self._execute_mutation_authenticated(
                mutation,
                variables,
                self.spark_user,
                self.endpoint_path,
            )

        assert result.data is not None
        assert result.data["approveRecap"]["success"] is True
        assert result.data["approveRecap"]["recap"]["approved"] is True
        assert mock_send.called

    @pytest.mark.asyncio
    async def test_approve_custom_recap_updates_approved_status(self):
        @sync_to_async
        def create_custom_recap():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Custom Sampling",
                slug="custom-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="Custom sampling recap",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            return recap_models.CustomRecap.objects.create(
                name="Custom recap",
                approved=False,
                event=self.event,
                tenant=self.tenant,
                custom_recap_template=template,
                created_by=self.spark_user,
            )

        custom_recap = await create_custom_recap()

        mutation = """
        mutation ApproveCustomRecap($input: ApproveCustomRecapInput!) {
            approveCustomRecap(input: $input) {
                success
                message
                customRecap {
                    id
                    approved
                }
            }
        }
        """
        variables = {
            "input": {
                "id": str(custom_recap.id),
                "approved": True,
            }
        }

        with patch("recaps.mutations.RecapApprovedNotificationMailer.send") as mock_send:
            result = await self._execute_mutation_authenticated(
                mutation,
                variables,
                self.spark_user,
                self.endpoint_path,
            )

        assert result.errors is None
        assert result.data is not None
        assert result.data["approveCustomRecap"]["success"] is True
        assert result.data["approveCustomRecap"]["customRecap"]["approved"] is True
        assert mock_send.called

        await sync_to_async(custom_recap.refresh_from_db)()
        assert custom_recap.approved is True
        assert custom_recap.updated_by_id == self.spark_user.id

    @pytest.mark.asyncio
    async def test_decline_custom_recap_updates_approved_status(self):
        @sync_to_async
        def create_custom_recap():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Decline Custom Sampling",
                slug="decline-custom-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="Decline custom sampling recap",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            return recap_models.CustomRecap.objects.create(
                name="Custom recap to decline",
                approved=True,
                event=self.event,
                tenant=self.tenant,
                custom_recap_template=template,
                created_by=self.spark_user,
            )

        custom_recap = await create_custom_recap()

        mutation = """
        mutation DeclineCustomRecap($input: DeclineCustomRecapInput!) {
            declineCustomRecap(input: $input) {
                success
                message
                customRecap {
                    id
                    approved
                }
            }
        }
        """
        variables = {
            "input": {
                "id": str(custom_recap.id),
            }
        }

        with patch("recaps.mutations.RecapApprovedNotificationMailer.send") as mock_send:
            result = await self._execute_mutation_authenticated(
                mutation,
                variables,
                self.spark_user,
                self.endpoint_path,
            )

        assert result.errors is None
        assert result.data is not None
        assert result.data["declineCustomRecap"]["success"] is True
        assert result.data["declineCustomRecap"]["customRecap"]["approved"] is False
        assert not mock_send.called

        await sync_to_async(custom_recap.refresh_from_db)()
        assert custom_recap.approved is False
        assert custom_recap.updated_by_id == self.spark_user.id

    @pytest.mark.asyncio
    async def test_generate_custom_recap_pdf_creates_custom_recap_file(self):
        @sync_to_async
        def create_custom_recap():
            system_user = self.get_system_user()
            file_type = self.create_file_type(name="PDF", extension=".pdf")
            event_type = EventType.objects.create(
                name="PDF Custom Sampling",
                slug="pdf-custom-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="PDF custom sampling recap",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            custom_recap = recap_models.CustomRecap.objects.create(
                name="Custom recap PDF",
                approved=True,
                event=self.event,
                tenant=self.tenant,
                custom_recap_template=template,
                created_by=self.spark_user,
            )
            return custom_recap, file_type

        custom_recap, file_type = await create_custom_recap()

        mutation = """
        mutation GenerateCustomRecapPdf($input: GenerateCustomRecapPdfInput!) {
            generateCustomRecapPdf(input: $input) {
                success
                message
                customRecapFile {
                    id
                    name
                    approved
                }
            }
        }
        """
        variables = {"input": {"id": str(custom_recap.id)}}

        with (
            patch("recaps.mutations.build_recap_pdf", return_value=b"pdf-bytes") as mock_build_pdf,
            patch("recaps.mutations.upload_bytes") as mock_upload_bytes,
        ):
            result = await self._execute_mutation_authenticated(
                mutation,
                variables,
                self.spark_user,
                self.endpoint_path,
            )

        assert result.errors is None
        assert result.data is not None
        assert result.data["generateCustomRecapPdf"]["success"] is True
        assert result.data["generateCustomRecapPdf"]["customRecapFile"]["approved"] is False
        assert result.data["generateCustomRecapPdf"]["customRecapFile"]["name"] == (
            "Custom Recap PDF - Custom recap PDF"
        )
        assert mock_build_pdf.called
        assert mock_upload_bytes.called

        @sync_to_async
        def get_custom_recap_file():
            return recap_models.CustomRecapFile.objects.get(
                custom_recap=custom_recap,
                file_type=file_type,
            )

        custom_recap_file = await get_custom_recap_file()
        assert str(custom_recap_file.url).startswith("recaps/pdfs/custom-")
        assert custom_recap_file.created_by_id == self.spark_user.id

    @pytest.mark.asyncio
    async def test_build_recaps_xlsx_supports_custom_recap(self):
        @sync_to_async
        def create_custom_recap_data():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Excel Custom Sampling",
                slug="excel-custom-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="Excel custom recap",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            field_type = recap_models.CustomRecapFieldType.objects.create(
                name="Text",
                created_by=system_user,
            )
            section = recap_models.RecapSection.objects.create(
                name="Summary",
                tenant=self.tenant,
                created_by=system_user,
            )
            custom_field = recap_models.CustomField.objects.create(
                name="Execution notes",
                required=False,
                custom_recap_template=template,
                custom_field_type=field_type,
                recap_section=section,
                created_by=system_user,
            )
            custom_recap = recap_models.CustomRecap.objects.create(
                name="Custom recap export",
                approved=True,
                event=self.event,
                tenant=self.tenant,
                custom_recap_template=template,
                ambassador=self.ambassador,
                created_by=self.spark_user,
            )
            product_type = ProductType.objects.create(
                name="Beverage",
                tenant=self.tenant,
                created_by=system_user,
            )
            product = Product.objects.create(
                name="Sample Product",
                product_type=product_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            type_of_good = recap_models.TypeOfGood.objects.create(
                name="Can",
                tenant=self.tenant,
                created_by=system_user,
            )
            recap_models.CustomRecapProductSample.objects.create(
                custom_recap=custom_recap,
                product=product,
                quantity=24,
                created_by=self.spark_user,
            )
            recap_models.CustomRecapSalePerformance.objects.create(
                custom_recap=custom_recap,
                product=product,
                type_of_good=type_of_good,
                price="12.5000",
                created_by=self.spark_user,
            )
            recap_models.CustomFieldValue.objects.create(
                custom_recap=custom_recap,
                custom_field=custom_field,
                value="Store manager asked for follow-up",
                created_by=self.spark_user,
            )
            file_type = self.create_file_type(name="Image", extension=".jpg")
            custom_recap_file = recap_models.CustomRecapFile.objects.create(
                name="Store photo",
                url="recap_files/custom/store-photo.jpg",
                custom_recap=custom_recap,
                file_type=file_type,
                created_by=self.spark_user,
            )
            return (
                recap_models.CustomRecap.objects.select_related(
                    "event",
                    "event__request__retailer",
                    "event__request__distributor",
                    "ambassador__user",
                )
                .prefetch_related(
                    "custom_recap_product_sample__product",
                    "custom_recap_sale_performance__product",
                    "custom_recap_sale_performance__type_of_good",
                    "custom_field_value__custom_field__custom_field_type",
                    "custom_field_value__custom_field__recap_section",
                    "custom_recap_files",
                )
                .get(id=custom_recap.id),
                str(custom_recap_file.uuid),
            )

        custom_recap, custom_recap_file_uuid = await create_custom_recap_data()
        workbook_bytes = build_recaps_xlsx(
            [custom_recap],
            frontend_base_url="https://spark-admin.example.com",
        )
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet_names = workbook.sheetnames

        recaps_rows = list(workbook["Recaps"].iter_rows(min_row=2, values_only=True))
        sample_rows = list(
            workbook["ProductSamples"].iter_rows(min_row=2, values_only=True)
        )
        sales_rows = list(
            workbook["SalesPerformance"].iter_rows(min_row=2, values_only=True)
        )
        file_rows = list(
            workbook["RecapFiles"].iter_rows(min_row=2, values_only=True)
        )
        file_hyperlink = workbook["RecapFiles"].cell(row=2, column=7).hyperlink
        section_headers = list(
            workbook["Summary"].iter_rows(min_row=1, max_row=1, values_only=True)
        )[0]
        section_rows = list(
            workbook["Summary"].iter_rows(min_row=2, values_only=True)
        )

        assert recaps_rows[0][1] == "Custom recap export"
        assert sample_rows[0][2] == "Sample Product"
        assert sales_rows[0][3] == "Can"
        assert file_rows[0][2] == "Store photo"
        assert file_hyperlink is not None
        assert (
            file_hyperlink.target
            == f"https://spark-admin.example.com/recap/file/{custom_recap_file_uuid}"
        )
        assert "ConsumerEngagements" not in sheet_names
        assert "ConsumerFeedback" not in sheet_names
        assert "AccountFeedback" not in sheet_names
        assert section_headers == ("recap_uuid", "recap_name", "Execution notes")
        assert section_rows[0][1] == "Custom recap export"
        assert section_rows[0][2] == "Store manager asked for follow-up"

    @pytest.mark.asyncio
    async def test_export_custom_recap_xlsx_returns_file_url(self):
        @sync_to_async
        def create_custom_recap():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Single Export Custom Sampling",
                slug="single-export-custom-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="Single export custom recap",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            return recap_models.CustomRecap.objects.create(
                name="Custom recap single export",
                approved=True,
                event=self.event,
                tenant=self.tenant,
                custom_recap_template=template,
                created_by=self.spark_user,
            )

        custom_recap = await create_custom_recap()

        mutation = """
        mutation ExportCustomRecapXlsx($input: ExportCustomRecapXlsxInput!) {
            exportCustomRecapXlsx(input: $input) {
                success
                message
                fileUrl
            }
        }
        """
        variables = {"input": {"id": str(custom_recap.id)}}

        class _FakeBlob:
            def __init__(self, name):
                self.name = name

            def delete(self):
                return None

        class _FakeBucket:
            def list_blobs(self, prefix):
                return [_FakeBlob(f"{prefix}old.xlsx")]

        class _FakeClient:
            def bucket(self, _name):
                return _FakeBucket()

        with (
            patch("recaps.mutations.get_gcs_client", return_value=_FakeClient()),
            patch("recaps.mutations.upload_bytes") as mock_upload_bytes,
            patch(
                "recaps.mutations.generate_download_url",
                return_value="https://example.com/custom-recap.xlsx",
            ),
        ):
            result = await self._execute_mutation_authenticated(
                mutation,
                variables,
                self.spark_user,
                self.endpoint_path,
            )

        assert result.errors is None
        assert result.data is not None
        assert result.data["exportCustomRecapXlsx"]["success"] is True
        assert (
            result.data["exportCustomRecapXlsx"]["fileUrl"]
            == "https://example.com/custom-recap.xlsx"
        )
        assert mock_upload_bytes.called

    @pytest.mark.asyncio
    async def test_recap_approved_mailer_template_renders(self):
        self.recap.approved = True
        await sync_to_async(self.recap.save)()
        recap = await sync_to_async(
            recap_models.Recap.objects.select_related(
                "event",
                "event__tenant",
                "job",
                "retailer",
                "timezone",
                "ambassador",
            ).get
        )(id=self.recap.id)

        mailer = RecapApprovedNotificationMailer(
            recap=recap,
            to_emails=[self.rmm_user.email],
            recipient_first_name=self.rmm_user.first_name,
            reply_to_email=self.rmm_user.email,
        )
        envelope = mailer.envelope()
        rendered_html = envelope.render_template()

        assert envelope.template == "recaps.templates.emails.recap_approved_notification"
        assert envelope.to_emails == [self.rmm_user.email]
        assert "Activation Summary" in rendered_html

    @pytest.mark.asyncio
    @override_settings(CLIENT_FRONTEND_URL="http://client.app")
    async def test_custom_recap_approved_mailer_template_renders(self):
        @sync_to_async
        def create_custom_recap():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Custom Approval",
                slug="custom-approval",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="Custom approval recap",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            return recap_models.CustomRecap.objects.create(
                name="Custom recap",
                approved=True,
                event=self.event,
                tenant=self.tenant,
                custom_recap_template=template,
                created_by=self.spark_user,
            )

        custom_recap = await create_custom_recap()
        custom_recap = await sync_to_async(
            recap_models.CustomRecap.objects.select_related(
                "event",
                "event__tenant",
                "job",
                "retailer",
                "timezone",
                "ambassador",
            ).get
        )(id=custom_recap.id)

        mailer = RecapApprovedNotificationMailer(
            recap=custom_recap,
            to_emails=[self.rmm_user.email],
            recipient_first_name=self.rmm_user.first_name,
            reply_to_email=self.rmm_user.email,
        )
        envelope = mailer.envelope()
        rendered_html = envelope.render_template()

        assert (
            envelope.template
            == "recaps.templates.emails.custom_recap_approved_notification"
        )
        assert (
            envelope.context["recap_link"]
            == f"http://client.app/recap/view-custom/{custom_recap.uuid}"
        )
        assert "Activation Summary" in rendered_html

    @pytest.mark.asyncio
    async def test_notify_recap_ready_for_review_sends_email_for_ambassador(self):
        with (
            override_settings(RECAP_REVIEW_COPY_EMAILS=["admin1@test.com", "admin2@test.com"]),
            patch("recaps.mutations.RecapReadyForReviewAdminMailer.send") as mock_send,
        ):
            await _notify_recap_ready_for_review_to_admins(
                recap=self.recap,
                created_by=self.ambassador_user,
            )
        assert mock_send.called

    @pytest.mark.asyncio
    async def test_notify_recap_ready_for_review_skips_non_ambassador(self):
        with (
            override_settings(RECAP_REVIEW_COPY_EMAILS=["admin1@test.com"]),
            patch("recaps.mutations.RecapReadyForReviewAdminMailer.send") as mock_send,
        ):
            await _notify_recap_ready_for_review_to_admins(
                recap=self.recap,
                created_by=self.spark_user,
            )
        assert not mock_send.called

    @pytest.mark.asyncio
    async def test_recap_query_returns_only_assigned_ambassador(self):
        query = """
        query Recap($uuid: ID!) {
            recap(uuid: $uuid) {
                id
                ambassador {
                    id
                }
            }
        }
        """
        variables = {"uuid": str(self.recap.uuid)}

        result = await self._execute_query_authenticated(
            query,
            variables,
            self.spark_user,
            self.endpoint_path,
        )

        assert result.errors is None
        assert result.data is not None
        assert result.data["recap"]["id"] == str(self.recap.id)
        assert result.data["recap"]["ambassador"] == {"id": str(self.ambassador.id)}

    @pytest.mark.asyncio
    async def test_recaps_query_filters_by_ambassador(self):
        other_recap = recap_models.Recap.objects.create(
            name="Other ambassador recap",
            approved=False,
            event=self.event,
            job=self.job,
            ambassador=self.other_ambassador,
            created_by=self.spark_user,
            updated_by=self.spark_user,
        )

        query = """
        query Recaps($filters: RecapFiltersInput) {
            recaps(filters: $filters) {
                totalCount
                edges {
                    node {
                        id
                        ambassador {
                            id
                        }
                    }
                }
            }
        }
        """
        variables = {"filters": {"ambassadorId": str(self.ambassador.id)}}

        result = await self._execute_query_authenticated(
            query,
            variables,
            self.spark_user,
            self.endpoint_path,
        )

        assert result.errors is None
        assert result.data is not None
        assert result.data["recaps"]["totalCount"] == 1
        assert result.data["recaps"]["edges"] == [
            {
                "node": {
                    "id": str(self.recap.id),
                    "ambassador": {"id": str(self.ambassador.id)},
                }
            }
        ]
        assert str(other_recap.id) not in {
            edge["node"]["id"] for edge in result.data["recaps"]["edges"]
        }

    @pytest.mark.asyncio
    async def test_recaps_query_filters_by_approved(self):
        approved_recap = recap_models.Recap.objects.create(
            name="Approved recap",
            approved=True,
            event=self.event,
            job=self.job,
            ambassador=self.other_ambassador,
            created_by=self.spark_user,
            updated_by=self.spark_user,
        )

        query = """
        query Recaps($filters: RecapFiltersInput) {
            recaps(filters: $filters) {
                totalCount
                edges {
                    node {
                        id
                        approved
                    }
                }
            }
        }
        """
        variables = {"filters": {"approved": True}}

        result = await self._execute_query_authenticated(
            query,
            variables,
            self.spark_user,
            self.endpoint_path,
        )

        assert result.errors is None
        assert result.data is not None
        assert result.data["recaps"]["totalCount"] == 1
        assert result.data["recaps"]["edges"] == [
            {
                "node": {
                    "id": str(approved_recap.id),
                    "approved": True,
                }
            }
        ]

    @pytest.mark.asyncio
    async def test_update_custom_recap_updates_existing_and_creates_new_field_values(self):
        @sync_to_async
        def create_custom_recap_data():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Sampling",
                slug="sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="Sampling recap",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            section = recap_models.RecapSection.objects.create(
                name="Main",
                tenant=self.tenant,
                created_by=system_user,
            )
            field_type = recap_models.CustomRecapFieldType.objects.create(
                name="Text",
                created_by=system_user,
            )
            existing_field = recap_models.CustomField.objects.create(
                name="Existing field",
                custom_recap_template=template,
                custom_field_type=field_type,
                recap_section=section,
                created_by=system_user,
            )
            new_field = recap_models.CustomField.objects.create(
                name="New field",
                custom_recap_template=template,
                custom_field_type=field_type,
                recap_section=section,
                created_by=system_user,
            )
            removed_field = recap_models.CustomField.objects.create(
                name="Removed field",
                custom_recap_template=template,
                custom_field_type=field_type,
                recap_section=section,
                created_by=system_user,
            )
            custom_recap = recap_models.CustomRecap.objects.create(
                name="Original custom recap",
                event=self.event,
                tenant=self.tenant,
                custom_recap_template=template,
                created_by=self.spark_user,
            )
            existing_value = recap_models.CustomFieldValue.objects.create(
                custom_recap=custom_recap,
                custom_field=existing_field,
                value="old value",
                created_by=self.spark_user,
            )
            removed_value = recap_models.CustomFieldValue.objects.create(
                custom_recap=custom_recap,
                custom_field=removed_field,
                value="remove me",
                created_by=self.spark_user,
            )

            return {
                "template_id": template.id,
                "custom_recap_id": custom_recap.id,
                "existing_value_id": existing_value.id,
                "existing_field_id": existing_field.id,
                "new_field_id": new_field.id,
                "removed_value_id": removed_value.id,
            }

        ids = await create_custom_recap_data()

        mutation = """
        mutation UpdateCustomRecap($input: UpdateCustomRecapInput!) {
            updateCustomRecap(input: $input) {
                success
                message
                customRecap {
                    id
                }
            }
        }
        """
        variables = {
            "input": {
                "id": str(ids["custom_recap_id"]),
                "name": "Updated custom recap",
                "eventId": str(self.event.id),
                "customRecapTemplateId": str(ids["template_id"]),
                "customFieldValues": [
                    {
                        "customFieldValueId": str(ids["existing_value_id"]),
                        "value": "updated value",
                    },
                    {
                        "customFieldId": str(ids["new_field_id"]),
                        "value": "new value",
                    },
                ],
            }
        }

        result = await self._execute_mutation_authenticated(
            mutation,
            variables,
            self.spark_user,
            self.endpoint_path,
        )

        assert result.errors is None
        assert result.data is not None
        assert result.data["updateCustomRecap"]["success"] is True

        @sync_to_async
        def get_field_values():
            return list(
                recap_models.CustomFieldValue.objects.filter(
                    custom_recap_id=ids["custom_recap_id"],
                )
                .order_by("custom_field_id")
                .values("id", "custom_field_id", "value", "updated_by_id")
            )

        values = await get_field_values()

        assert {
            (item["custom_field_id"], item["value"]) for item in values
        } == {
            (ids["existing_field_id"], "updated value"),
            (ids["new_field_id"], "new value"),
        }
        assert any(
            item["id"] == ids["existing_value_id"]
            and item["updated_by_id"] == self.spark_user.id
            for item in values
        )
        assert ids["removed_value_id"] not in {item["id"] for item in values}

    @pytest.mark.asyncio
    async def test_create_custom_recap_template_creates_custom_fields(self):
        @sync_to_async
        def create_template_dependencies():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Field Template Sampling",
                slug="field-template-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            section = recap_models.RecapSection.objects.create(
                name="Main",
                tenant=self.tenant,
                created_by=system_user,
            )
            field_type = recap_models.CustomRecapFieldType.objects.create(
                name="Text",
                created_by=system_user,
            )

            return {
                "event_type_id": event_type.id,
                "section_id": section.id,
                "field_type_id": field_type.id,
            }

        ids = await create_template_dependencies()

        mutation = """
        mutation CreateCustomRecapTemplate($input: CreateCustomRecapTemplateInput!) {
            createCustomRecapTemplate(input: $input) {
                success
                message
                customRecapTemplate {
                    id
                }
            }
        }
        """
        variables = {
            "input": {
                "name": "Template with fields",
                "eventTypeId": str(ids["event_type_id"]),
                "customFields": [
                    {
                        "name": "Display notes",
                        "customFieldTypeId": str(ids["field_type_id"]),
                        "recapSectionId": str(ids["section_id"]),
                        "required": True,
                    }
                ],
            }
        }

        result = await self._execute_mutation_authenticated(
            mutation,
            variables,
            self.spark_user,
            self.endpoint_path,
        )

        assert result.errors is None
        assert result.data is not None
        assert result.data["createCustomRecapTemplate"]["success"] is True

        template_id = int(
            result.data["createCustomRecapTemplate"]["customRecapTemplate"]["id"]
        )

        @sync_to_async
        def get_custom_fields():
            return list(
                recap_models.CustomField.objects.filter(
                    custom_recap_template_id=template_id,
                ).values(
                    "name",
                    "custom_field_type_id",
                    "recap_section_id",
                    "required",
                    "created_by_id",
                )
            )

        custom_fields = await get_custom_fields()

        assert custom_fields == [
            {
                "name": "Display notes",
                "custom_field_type_id": ids["field_type_id"],
                "recap_section_id": ids["section_id"],
                "required": True,
                "created_by_id": self.spark_user.id,
            }
        ]

    @pytest.mark.asyncio
    async def test_update_custom_recap_template_syncs_custom_fields(self):
        @sync_to_async
        def create_template_with_fields():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Update Field Template Sampling",
                slug="update-field-template-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="Original template",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            section = recap_models.RecapSection.objects.create(
                name="Main update",
                tenant=self.tenant,
                created_by=system_user,
            )
            field_type = recap_models.CustomRecapFieldType.objects.create(
                name="Text update",
                created_by=system_user,
            )
            existing_field = recap_models.CustomField.objects.create(
                name="Existing field",
                custom_recap_template=template,
                custom_field_type=field_type,
                recap_section=section,
                required=False,
                created_by=system_user,
            )
            removed_field = recap_models.CustomField.objects.create(
                name="Removed field",
                custom_recap_template=template,
                custom_field_type=field_type,
                recap_section=section,
                created_by=system_user,
            )

            return {
                "event_type_id": event_type.id,
                "template_id": template.id,
                "section_id": section.id,
                "field_type_id": field_type.id,
                "existing_field_id": existing_field.id,
                "removed_field_id": removed_field.id,
            }

        ids = await create_template_with_fields()

        mutation = """
        mutation UpdateCustomRecapTemplate($input: UpdateCustomRecapTemplateInput!) {
            updateCustomRecapTemplate(input: $input) {
                success
                message
                customRecapTemplate {
                    id
                }
            }
        }
        """
        variables = {
            "input": {
                "id": str(ids["template_id"]),
                "name": "Updated template",
                "eventTypeId": str(ids["event_type_id"]),
                "customFields": [
                    {
                        "id": str(ids["existing_field_id"]),
                        "name": "Updated existing field",
                        "customFieldTypeId": str(ids["field_type_id"]),
                        "recapSectionId": str(ids["section_id"]),
                        "required": True,
                    },
                    {
                        "name": "Created field",
                        "customFieldTypeId": str(ids["field_type_id"]),
                        "recapSectionId": str(ids["section_id"]),
                        "required": False,
                    },
                ],
            }
        }

        result = await self._execute_mutation_authenticated(
            mutation,
            variables,
            self.spark_user,
            self.endpoint_path,
        )

        assert result.errors is None
        assert result.data is not None
        assert result.data["updateCustomRecapTemplate"]["success"] is True

        @sync_to_async
        def get_custom_fields():
            return list(
                recap_models.CustomField.objects.filter(
                    custom_recap_template_id=ids["template_id"],
                )
                .order_by("name")
                .values("id", "name", "required", "updated_by_id", "created_by_id")
            )

        custom_fields = await get_custom_fields()

        assert {field["name"] for field in custom_fields} == {
            "Created field",
            "Updated existing field",
        }
        assert ids["removed_field_id"] not in {field["id"] for field in custom_fields}
        assert any(
            field["id"] == ids["existing_field_id"]
            and field["required"] is True
            and field["updated_by_id"] == self.spark_user.id
            for field in custom_fields
        )
        assert any(
            field["name"] == "Created field"
            and field["created_by_id"] == self.spark_user.id
            for field in custom_fields
        )

    @pytest.mark.asyncio
    async def test_custom_recap_templates_filters_by_tenant_and_event_type(self):
        @sync_to_async
        def create_template_filter_data():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Template Filter Sampling",
                slug="template-filter-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            other_event_type = EventType.objects.create(
                name="Other Template Filter Sampling",
                slug="other-template-filter-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            matching_template = recap_models.CustomRecapTemplate.objects.create(
                name="Matching template",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            recap_models.CustomRecapTemplate.objects.create(
                name="Other template",
                event_type=other_event_type,
                tenant=self.tenant,
                created_by=system_user,
            )

            return {
                "event_type_id": event_type.id,
                "matching_template_id": matching_template.id,
            }

        ids = await create_template_filter_data()

        query = """
        query CustomRecapTemplates($filters: CustomRecapTemplateFiltersInput) {
            customRecapTemplates(filters: $filters) {
                totalCount
                edges {
                    node {
                        id
                        name
                        tenant {
                            id
                            name
                        }
                        eventType {
                            id
                            name
                        }
                    }
                }
            }
        }
        """
        variables = {
            "filters": {
                "tenantId": str(self.tenant.id),
                "eventTypeId": str(ids["event_type_id"]),
            }
        }

        result = await self._execute_query_authenticated(
            query,
            variables,
            self.spark_user,
            self.endpoint_path,
        )

        assert result.errors is None
        assert result.data is not None
        assert result.data["customRecapTemplates"] == {
            "totalCount": 1,
            "edges": [
                {
                    "node": {
                        "id": str(ids["matching_template_id"]),
                        "name": "Matching template",
                        "tenant": {
                            "id": str(self.tenant.id),
                            "name": self.tenant.name,
                        },
                        "eventType": {
                            "id": str(ids["event_type_id"]),
                            "name": "Template Filter Sampling",
                        },
                    }
                },
            ],
        }

    @pytest.mark.asyncio
    async def test_custom_recap_template_returns_tenant_and_event_type(self):
        @sync_to_async
        def create_template():
            system_user = self.get_system_user()
            event_type = EventType.objects.create(
                name="Single Template Sampling",
                slug="single-template-sampling",
                tenant=self.tenant,
                created_by=system_user,
            )
            template = recap_models.CustomRecapTemplate.objects.create(
                name="Single template",
                event_type=event_type,
                tenant=self.tenant,
                created_by=system_user,
            )
            return {
                "template_id": template.id,
                "event_type_id": event_type.id,
            }

        ids = await create_template()

        query = """
        query CustomRecapTemplate($id: ID!) {
            customRecapTemplate(id: $id) {
                id
                name
                tenant {
                    id
                    name
                }
                eventType {
                    id
                    name
                }
            }
        }
        """
        variables = {"id": str(ids["template_id"])}

        result = await self._execute_query_authenticated(
            query,
            variables,
            self.spark_user,
            self.endpoint_path,
        )

        assert result.errors is None
        assert result.data is not None
        assert result.data["customRecapTemplate"] == {
            "id": str(ids["template_id"]),
            "name": "Single template",
            "tenant": {
                "id": str(self.tenant.id),
                "name": self.tenant.name,
            },
            "eventType": {
                "id": str(ids["event_type_id"]),
                "name": "Single Template Sampling",
            },
        }


@pytest.mark.django_db(transaction=True)
class TestBaRecapEditGuards(JobsGraphQLTestCase):
    """Authorization guards for BA-driven custom-recap edits.

    A Brand Ambassador may edit their OWN custom recap, ONLY until it's
    approved, and may NEVER set the `approved` flag — regardless of which
    mutation/input variant they use (both updateCustomRecapMobile and the
    web updateCustomRecap are exposed to the BA app via
    RecapMutationsMobile). Admins/clients are unaffected.
    """

    UPDATE_MOBILE = """
    mutation UpdateCustomRecapMobile($input: UpdateCustomRecapMobileInput!) {
        updateCustomRecapMobile(input: $input) {
            success
            message
            customRecap { id approved }
        }
    }
    """

    UPDATE_WEB = """
    mutation UpdateCustomRecap($input: UpdateCustomRecapInput!) {
        updateCustomRecap(input: $input) {
            success
            message
            customRecap { id approved }
        }
    }
    """

    @pytest.fixture(autouse=True)
    def setup(self, db):
        from config.schema_mobile import schema_mobile
        from config.schema_spark import schema_spark

        self.roles = self.setup_default_roles()
        self.tenant = self.create_tenant(name="Edit Guard Tenant")

        self.spark_user = self.create_user(
            username="edit_spark@test.com",
            email="edit_spark@test.com",
            role=self.roles["spark_admin"],
            password="testpass123",
        )
        self.create_tenanted_user(user=self.spark_user, tenant=self.tenant)

        self.owner_user = self.create_user(
            username="edit_owner_amb@test.com",
            email="edit_owner_amb@test.com",
            role=self.roles["ambassador"],
            password="testpass123",
        )
        self.owner_ambassador = self.create_ambassador(user=self.owner_user)
        self.create_tenanted_user(user=self.owner_user, tenant=self.tenant)

        self.other_user = self.create_user(
            username="edit_other_amb@test.com",
            email="edit_other_amb@test.com",
            role=self.roles["ambassador"],
            password="testpass123",
        )
        self.other_ambassador = self.create_ambassador(user=self.other_user)
        self.create_tenanted_user(user=self.other_user, tenant=self.tenant)

        self.event = self.create_event(
            name="Edit Guard Event",
            tenant=self.tenant,
            address="1 Edit St",
        )

        system_user = self.get_system_user()
        event_type = EventType.objects.create(
            name="Edit Guard Sampling",
            slug="edit-guard-sampling",
            tenant=self.tenant,
            created_by=system_user,
        )
        template = recap_models.CustomRecapTemplate.objects.create(
            name="Edit guard recap",
            event_type=event_type,
            tenant=self.tenant,
            created_by=system_user,
        )
        section = recap_models.RecapSection.objects.create(
            name="Main",
            tenant=self.tenant,
            created_by=system_user,
        )
        field_type = recap_models.CustomRecapFieldType.objects.create(
            name="Text",
            created_by=system_user,
        )
        self.custom_field = recap_models.CustomField.objects.create(
            name="Notes",
            custom_recap_template=template,
            custom_field_type=field_type,
            recap_section=section,
            created_by=system_user,
        )
        self.template = template
        self.recap = recap_models.CustomRecap.objects.create(
            name="Owner's recap",
            approved=False,
            event=self.event,
            tenant=self.tenant,
            ambassador=self.owner_ambassador,
            custom_recap_template=template,
            created_by=self.owner_user,
        )
        recap_models.CustomFieldValue.objects.create(
            custom_recap=self.recap,
            custom_field=self.custom_field,
            value="original",
            created_by=self.owner_user,
        )

        self.schema = schema_mobile
        self.spark_schema = schema_spark
        self.mobile_path = "/api/v1/graphql/mobile"
        self.spark_path = "/api/v1/graphql/spark"

    @sync_to_async
    def _refresh(self):
        self.recap.refresh_from_db()
        return self.recap

    @sync_to_async
    def _notes_value(self):
        return (
            recap_models.CustomFieldValue.objects.filter(
                custom_recap=self.recap, custom_field=self.custom_field
            )
            .values_list("value", flat=True)
            .first()
        )

    @pytest.mark.asyncio
    async def test_ba_can_edit_own_unapproved_recap(self):
        variables = {
            "input": {
                "id": str(self.recap.id),
                "name": "Owner's recap",
                "customFieldValues": [
                    {"customFieldId": str(self.custom_field.id), "value": "edited"},
                ],
            }
        }
        result = await self._execute_mutation_authenticated(
            self.UPDATE_MOBILE, variables, self.owner_user, self.mobile_path
        )
        assert result.errors is None
        assert result.data["updateCustomRecapMobile"]["success"] is True
        assert await self._notes_value() == "edited"

    @pytest.mark.asyncio
    async def test_ba_cannot_edit_another_bas_recap_mobile(self):
        variables = {
            "input": {
                "id": str(self.recap.id),
                "name": "Hijacked",
                "customFieldValues": [
                    {"customFieldId": str(self.custom_field.id), "value": "hijack"},
                ],
            }
        }
        result = await self._execute_mutation_authenticated(
            self.UPDATE_MOBILE, variables, self.other_user, self.mobile_path
        )
        assert result.data["updateCustomRecapMobile"]["success"] is False
        assert "not found" in result.data["updateCustomRecapMobile"]["message"].lower()
        assert await self._notes_value() == "original"

    @pytest.mark.asyncio
    async def test_ba_cannot_edit_approved_recap(self):
        @sync_to_async
        def approve():
            self.recap.approved = True
            self.recap.save(update_fields=["approved"])

        await approve()
        variables = {
            "input": {
                "id": str(self.recap.id),
                "name": "Owner's recap",
                "customFieldValues": [
                    {"customFieldId": str(self.custom_field.id), "value": "late edit"},
                ],
            }
        }
        result = await self._execute_mutation_authenticated(
            self.UPDATE_MOBILE, variables, self.owner_user, self.mobile_path
        )
        assert result.data["updateCustomRecapMobile"]["success"] is False
        assert "approved" in result.data["updateCustomRecapMobile"]["message"].lower()
        assert await self._notes_value() == "original"

    @pytest.mark.asyncio
    async def test_ba_cannot_self_approve_via_web_input(self):
        # A BA hitting the web-input updateCustomRecap (also exposed to the
        # app) with approved:true — the edit is allowed but approval is
        # silently ignored (caller-role gate).
        variables = {
            "input": {
                "id": str(self.recap.id),
                "name": "Owner's recap",
                "eventId": str(self.event.id),
                "customRecapTemplateId": str(self.template.id),
                "approved": True,
                "customFieldValues": [
                    {"customFieldId": str(self.custom_field.id), "value": "edited"},
                ],
            }
        }
        result = await self._execute_mutation_authenticated(
            self.UPDATE_WEB, variables, self.owner_user, self.mobile_path
        )
        assert result.errors is None
        assert result.data["updateCustomRecap"]["success"] is True
        recap = await self._refresh()
        assert recap.approved is False

    @pytest.mark.asyncio
    async def test_ba_cannot_edit_another_bas_recap_via_web_input(self):
        variables = {
            "input": {
                "id": str(self.recap.id),
                "name": "Hijacked",
                "eventId": str(self.event.id),
                "customRecapTemplateId": str(self.template.id),
                "customFieldValues": [
                    {"customFieldId": str(self.custom_field.id), "value": "hijack"},
                ],
            }
        }
        result = await self._execute_mutation_authenticated(
            self.UPDATE_WEB, variables, self.other_user, self.mobile_path
        )
        assert result.data["updateCustomRecap"]["success"] is False
        assert "not found" in result.data["updateCustomRecap"]["message"].lower()
        assert await self._notes_value() == "original"

    @pytest.mark.asyncio
    async def test_admin_can_approve_via_update(self):
        # Admins are unaffected by the BA gate — approval via update works.
        variables = {
            "input": {
                "id": str(self.recap.id),
                "name": "Owner's recap",
                "eventId": str(self.event.id),
                "customRecapTemplateId": str(self.template.id),
                "approved": True,
                "customFieldValues": [
                    {"customFieldId": str(self.custom_field.id), "value": "original"},
                ],
            }
        }
        self.schema = self.spark_schema
        result = await self._execute_mutation_authenticated(
            self.UPDATE_WEB, variables, self.spark_user, self.spark_path
        )
        assert result.errors is None
        assert result.data["updateCustomRecap"]["success"] is True
        recap = await self._refresh()
        assert recap.approved is True
